from django.shortcuts import render
from ..forms import SistemaGPIparaContadoresForm
from ..models import SistemaGPIparaContadores
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import time
from io import BytesIO

def cadastro_sistema_GPI_contadores(request):
    if request.method == 'POST':
        form = SistemaGPIparaContadoresForm(request.POST)
        if form.is_valid():
            # Verifica se já existe inscrição para a mesma matrícula em qualquer turma
            if SistemaGPIparaContadores.objects.filter(cpf=form.cleaned_data['cpf']).exists():
                context = {
                    'form': form,
                    'titulo': 'Treinamento do Sistema GPI para Contadores',
                    'subtitulo': 'Inscrição',
                    'mensagem': '<span class="text-danger">Esta matrícula já está inscrita em uma das turmas.</span>'
                }
                return render(request, 'formfacil/formfacil_form.html', context)
            # Verifica se já o limite de vagas ja foi atingido
            if SistemaGPIparaContadores.objects.filter(turma=form.cleaned_data['turma']).count() >= 30:
                context = {
                    'form': form,
                    'titulo': 'Treinamento do Sistema GPI para Contadores',
                    'subtitulo': 'Inscrição',
                    'mensagem': '<span class="text-danger">Limite de vagas atingido para esta turma.</span>'
                }
                return render(request, 'formfacil/formfacil_form.html', context)
            form.save()
            context = {
                'titulo': 'Treinamento do Sistema GPI para Contadores',
                'subtitulo': 'Inscrição',
                'mensagem': "<span class='text-success'><i class='fa-solid fa-circle-check me-2'></i>Formulário enviado com sucesso!</span>"
            }
            return render(request, 'formfacil/formfacil_success.html', context)
        else:
            context = {
                    'form': form,
                    'titulo': 'Treinamento do Sistema GPI para Contadores',
                    'subtitulo': 'Inscrição',
                    'mensagem': '<span class="text-danger">Erro ao enviar formulario.</span>'
                }
            return render(request, 'formfacil/formfacil_form.html', context)
    else:
        form = SistemaGPIparaContadoresForm()
        context = {
            'form': form,
            'titulo': 'Treinamento do Sistema GPI para Contadores',
            'subtitulo': 'Inscrição',
            'mensagem': (
                'Turma 1: 10h às 12h | 04/07/2025<br>'
                'Turma 2: 14h às 16h | 04/07/2025<br>'
                'Local: Sala de treinamento da secretaria de ciência e tecnologia, inovação e desenvolvimento econômico - Av. Alberto Braune, 223 - 2º andar.'
            )
        }
        return render(request, 'formfacil/formfacil_form.html', context)

def visualizar_sistema_GPI_contadores(request):
    turmas = dict(SistemaGPIparaContadores.TURMAS)
    registros = {turma: SistemaGPIparaContadores.objects.filter(turma=key) for key, turma in turmas.items()}
    return render(request, 'formfacil/visualizar_sistema_GPI_contadores.html', {'registros': registros, 'turmas': turmas})

def exportar_sistema_GPI_contadores(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Treinamento do Sistema GPI para Contadores'
    cadastros = SistemaGPIparaContadores.objects.all().order_by('turma', 'nome')
    ws.append(['Nome', 'CPF', 'Telefone', 'Empresa', 'turma', 'Data de inscrição'])
    for item in cadastros:
        ws.append([
            item.nome,
            item.cpf,
            item.telefone,
            item.empresa,
            item.turma,
            item.dt_inscricao.strftime('%d/%m/%Y %H:%M:%S')
        ])
    col_widths = [30, 20, 25, 25, 20, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    total_rows = ws.max_row
    total_cols = ws.max_column
    first_col = get_column_letter(1)
    last_col = get_column_letter(total_cols)
    table_ref = f"{first_col}1:{last_col}{total_rows}"
    timestamp = int(time.time())
    table_name = f"ProcDigital_{timestamp}"
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Inscriçoes_sistema_GPI_contadores.xlsx'
    response.write(output.getvalue())
    return response
