from django.shortcuts import render
from ..forms import *
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.utils import timezone
from django.db.models import IntegerField
from django.db.models.functions import Cast

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from django.http import HttpResponse
from datetime import datetime

def cadastroTreinamentoTributarioEmissoresTaxas(request):
    if request.method == 'POST':
        form = CadastroAulasEmissoresForm(request.POST)
        if form.is_valid():
            form.save()
            context = {
                'titulo': 'TREINAMENTO TRIBUTÁRIO - EMISSORES DE TAXAS',    
                'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
                'mensagem': "<span class='text-success'><i class='fa-solid fa-circle-check me-2'></i>Formulário enviado com sucesso!</span>"
            }
            return render(request, 'formfacil/formfacil_success.html', context)
    else:
        form = CadastroAulasEmissoresForm()
        
    context = {
            'form': form,
            'titulo': 'TREINAMENTO TRIBUTÁRIO - EMISSORES DE TAXAS',    
            'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
            'mensagem': (
                "Apenas para as secretarias: Smomu, Serv. Públicos, Meio Ambiente e Saúde (Vigilância Sanitária). <br><br>"
                "Data: 07/10 (segunda-feira) <br>"
                "Horario : das 10h às 12h. <br>"
                "Local: Acianf - Av. Alberto Braune, 111 (ao lado da loja Ortobom)"
            )
        }
    return render(request, 'formfacil/formfacil_form.html', context)

def cadastroTreinamentoTributarioContadores(request):
    qntd = Cadastro_Aulas_Treinamento_Tributario_Contadores.objects.all().count()
    if qntd >= 82:
        context = {
            'form': 'Excedeu o limite de inscrições para este treinamento.',
            'titulo': 'TREINAMENTO TRIBUTÁRIO - CONTADORES',    
            'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
            'mensagem': (
                "Data: 11/10 (sexta-feira) <br>"
                "Horario : das 10h às 12h. <br>"
                "Local: SEBRAE (ARP)"
            )
        }
        return render(request, 'formfacil/formfacil_form.html', context)
    if request.method == 'POST':       
        form = CadastroAulasContadoresForm(request.POST)
        if form.is_valid():
            form.save()
            context = {
                'titulo': 'TREINAMENTO TRIBUTÁRIO - CONTADORES',   
                'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
                'mensagem': "<span class='text-success'><i class='fa-solid fa-circle-check me-2'></i>Formulário enviado com sucesso!</span>"
            }
            return render(request, 'formfacil/formfacil_success.html', context)
    else:
        form = CadastroAulasContadoresForm()
        
    context = {
            'form': form,
            'titulo': 'TREINAMENTO TRIBUTÁRIO - CONTADORES',    
            'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
            'mensagem': (
                "Data: 11/10 (sexta-feira) <br>"
                "Horario : das 10h às 12h. <br>"
                "Local: SEBRAE (ARP)"
            )
        }
    return render(request, 'formfacil/formfacil_form.html', context)

def visualizarDados_TT_Emissores(request):
    cadastros_emissores = Cadastro_Aulas_Treinamento_Tributario_Emissores_Taxas.objects.all()

    context = {
        'cadastros_emissores': cadastros_emissores
    }
    return render(request, 'formfacil/template_emissores.html', context)

def visualizarDados_Decreto_Portaria_Atos_Prefeito(request):
    registros = []
    for horario in Inscricao_Decretos_Portaria_E_Atos_Do_Prefeito.HORARIOS_CHOICES:
        itens = Inscricao_Decretos_Portaria_E_Atos_Do_Prefeito.objects.filter(horarios=horario[1])
        registros.append({
            'horario': horario[1],
            'itens': itens,
            'total': itens.count(),
        })
    context = {
        'registros': registros
    }
    return render(request, 'formfacil/template_atos_do_prefeito.html', context)


def visualizarDados_TT_Contadores(request):
    cadastros_contadores = Cadastro_Aulas_Treinamento_Tributario_Contadores.objects.all()
    context = {
        'cadastros_contadores': cadastros_contadores
    }
    return render(request, 'formfacil/template_contadores.html', context)


def exportar_aulas_tributario_emissores_to_excel(request):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'TREINAMENTO TRIBUTÁRIO - EMISSORES DE TAXAS'

    # Load data from the database
    cadastros = Cadastro_Aulas_Treinamento_Tributario_Emissores_Taxas.objects.all()
    ws.append(['Nome', 'CPF', 'Matrícula', 'Secretaria', 'Setor', 'Telefone', 'Data de inclusão'])
    # Write the data to the worksheet
    for item in cadastros:
        ws.append([
            item.nome,
            item.cpf,
            item.matricula,
            item.secretaria,
            item.setor,            
            item.telefone,
            item.dt_registro.astimezone().strftime('%d/%m/%Y %H:%M:%S')
        ])
    # for r in dataframe_to_rows(df, index=False, header=True):
        # ws.append(r)

    # Create a table
    table = Table(displayName=f"TreinamentoTributarioEmissoresTaxas", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the workbook to a BytesIO object
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set up the response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=TreinamentoTributarioEmissoresTaxas.xlsx'
    response.write(output.getvalue())
    return response


def exportar_aulas_tributario_contadores_to_excel(request):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'TREINAMENTO TRIBUTÁRIO - CONTADORES'

    # Load data from the database
    cadastros = Cadastro_Aulas_Treinamento_Tributario_Contadores.objects.all()
    ws.append(['Nome', 'CPF', 'Matrícula', 'Secretaria', 'Setor', 'Telefone', 'Data de inclusão'])
    # Write the data to the worksheet
    for item in cadastros:
        ws.append([
            item.nome,
            item.cpf,
            item.matricula,
            item.secretaria,
            item.setor,            
            item.telefone,
            item.dt_registro.astimezone().strftime('%d/%m/%Y %H:%M:%S')
        ])
    # for r in dataframe_to_rows(df, index=False, header=True):
        # ws.append(r)

    # Create a table
    table = Table(displayName=f"TreinamentoTributarioContadores", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the workbook to a BytesIO object
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set up the response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=TreinamentoTributarioContadores.xlsx'
    response.write(output.getvalue())
    return response


def logCadastrosRepetidosTTEmissores(request):
    # Obtém as matrículas que têm registros repetidos a partir da data especificada
    matriculas_repetidas = (
        cadastroTreinamentoTributarioEmissoresTaxas.objects.filter(~Q(matricula='')) 
        .values('matricula')  # Agrupa por matrícula
        .annotate(total=Count('matricula'))  # Conta quantas vezes cada matrícula aparece
        .filter(total__gt=1)  # Filtra apenas as matrículas que aparecem mais de uma vez
        .values_list('matricula', flat=True)  # Obtém uma lista de matrículas
    )

    # Obtém os nomes que têm registros repetidos a partir da data especificada (quando não têm matrícula)
    nomes_repetidos = (
        cadastroTreinamentoTributarioEmissoresTaxas.objects.filter(Q(matricula=''))  # Filtra apenas os que não têm matrícula
        .values('nome')  # Agrupa por nome
        .annotate(total=Count('nome'))  # Conta quantas vezes cada nome aparece
        .filter(total__gt=1)  # Filtra apenas os nomes que aparecem mais de uma vez
        .values_list('nome', flat=True)  # Obtém uma lista de nomes
    )

    # Obtém todos os registros para as matrículas repetidas ou nomes repetidos a partir da data especificada
    cadastros_repetidos = cadastroTreinamentoTributarioEmissoresTaxas.objects.filter(
        Q(matricula__in=matriculas_repetidas) | Q(nome__in=nomes_repetidos)
    )

    # Renderiza o template com os cadastros repetidos
    return render(request, 'formfacil/cadastrosrepetidos.html', {'cadastros': cadastros_repetidos, 'total': cadastros_repetidos.count()})

def logCadastrosRepetidosTTContadores(request):
    # Data específica para filtrar os registros
    data_especifica = timezone.datetime(2024, 9, 9)
    
    # Obtém as matrículas que têm registros repetidos a partir da data especificada
    matriculas_repetidas = (
        cadastroTreinamentoTributarioContadores.objects
        .filter(dt_registro__gte=data_especifica)  # Filtra registros a partir da data especificada
        .filter(~Q(matricula=''))  # Filtra apenas os que têm matrícula
        .values('matricula')  # Agrupa por matrícula
        .annotate(total=Count('matricula'))  # Conta quantas vezes cada matrícula aparece
        .filter(total__gt=1)  # Filtra apenas as matrículas que aparecem mais de uma vez
        .values_list('matricula', flat=True)  # Obtém uma lista de matrículas
    )

    # Obtém os nomes que têm registros repetidos a partir da data especificada (quando não têm matrícula)
    nomes_repetidos = (
        cadastroTreinamentoTributarioContadores.objects
        .filter(dt_registro__gte=data_especifica)  # Filtra registros a partir da data especificada
        .filter(Q(matricula=''))  # Filtra apenas os que não têm matrícula
        .values('nome')  # Agrupa por nome
        .annotate(total=Count('nome'))  # Conta quantas vezes cada nome aparece
        .filter(total__gt=1)  # Filtra apenas os nomes que aparecem mais de uma vez
        .values_list('nome', flat=True)  # Obtém uma lista de nomes
    )

    # Obtém todos os registros para as matrículas repetidas ou nomes repetidos a partir da data especificada
    cadastros_repetidos = cadastroTreinamentoTributarioContadores.objects.filter(
        Q(matricula__in=matriculas_repetidas) | Q(nome__in=nomes_repetidos),
    )

    # Renderiza o template com os cadastros repetidos
    return render(request, 'formfacil/cadastrosrepetidos.html', {'cadastros': cadastros_repetidos, 'total': cadastros_repetidos.count()})


from django.contrib import messages
def cadastroDecretos2024(request):
    
    if request.method == 'POST':
        form = CadastroDecretos2024Form(request.POST)
        if form.is_valid():
            inscricao = form.save(commit=False)
            horarios_disponiveis = Inscricao_Decretos_Portaria_E_Atos_Do_Prefeito.get_qnt_inscritos()
            print(horarios_disponiveis)
            if not horarios_disponiveis.get(inscricao.horarios, False):
                messages.error(request, f"Não há vagas disponíveis para o horário <strong>{inscricao.horarios}</strong>.")
            else:
                # Salva a inscrição
                inscricao.save()                            
                context = {
                    'titulo': 'Decretos e Portaria Atos do Prefeito',    
                    'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
                    'mensagem': "<span class='text-success'><i class='fa-solid fa-circle-check me-2'></i>Formulário enviado com sucesso!</span>"
                }
                return render(request, 'formfacil/formfacil_success.html', context)
    else:
        form = CadastroDecretos2024Form()
        
    context = {
            'form': form,
            'titulo': 'Decretos e Portaria Atos do Prefeito',       
            'subtitulo': 'Subsecretaria de Tecnologia da Informação e Comunicação',
            'mensagem': (                
                "Data: 12/12 (Quinta-feira) <br>"
                "Horário : das 14h às 16h. <br>"
                "Local: À definir"
            )
        }
    return render(request, 'formfacil/formfacil_form.html', context)

def exportar_decretos_portaria_atos_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Decretos e Portaria Atos do Prefeito'
    cadastros = Inscricao_Decretos_Portaria_E_Atos_Do_Prefeito.objects.all().order_by('horarios', 'nome')
    ws.append(['Nome', 'Matrícula', 'Secretaria', 'Setor', 'Telefone', 'Turma', 'Dt. Registro'])
    for item in cadastros:
        ws.append([
            item.nome,
            item.matricula,
            item.secretaria,
            item.setor,
            item.telefone,
            item.horarios,
            item.dt_registro.astimezone().strftime('%d/%m/%Y %H:%M:%S')
        ])
    # Ajusta a largura das colunas
    from openpyxl.utils import get_column_letter
    import time
    col_widths = [25, 15, 20, 20, 15, 15, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    # Corrige o intervalo da tabela
    total_rows = ws.max_row
    total_cols = ws.max_column
    first_col = get_column_letter(1)
    last_col = get_column_letter(total_cols)
    table_ref = f"{first_col}1:{last_col}{total_rows}"
    # Gera um nome único para a tabela
    timestamp = int(time.time())
    table_name = f"DecretosAtos_{timestamp}"
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=DecretosPortariaAtos.xlsx'
    response.write(output.getvalue())
    return response