from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import time
from ..forms import PadronizacaoPagamentoInscricaoForm
from ..models import PadronizacaoPagamentoInscricao

def cadastro_padronizacao_pagamento(request):
    if request.method == 'POST':
        form = PadronizacaoPagamentoInscricaoForm(request.POST)
        if form.is_valid():
            # Verifica se já existe inscrição para a mesma matrícula em qualquer turma
            if PadronizacaoPagamentoInscricao.objects.filter(matricula=form.cleaned_data['matricula']).exists():
                context = {
                    'form': form,
                    'titulo': 'Padronização do Processo de Pagamento',
                    'subtitulo': 'Inscrição',
                    'mensagem': '<span class="text-danger">Esta matrícula já está inscrita em uma das turmas.</span>'
                }
                return render(request, 'formfacil/formfacil_form.html', context)
            if PadronizacaoPagamentoInscricao.objects.filter(turma=form.cleaned_data['turma']).count() >= 28:
                context = {
                    'form': form,
                    'titulo': 'Padronização do Processo de Pagamento',
                    'subtitulo': 'Inscrição',
                    'mensagem': '<span class="text-danger">Limite de vagas atingido para esta turma.</span>'
                }
                return render(request, 'formfacil/formfacil_form.html', context)
            form.save()
            context = {
                'titulo': 'Padronização do Processo de Pagamento',
                'subtitulo': 'Inscrição',
                'mensagem': "<span class='text-success'><i class='fa-solid fa-circle-check me-2'></i>Formulário enviado com sucesso!</span>"
            }
            return render(request, 'formfacil/formfacil_success.html', context)
    else:
        form = PadronizacaoPagamentoInscricaoForm()
    context = {
        'form': form,
        'titulo': 'Padronização do Processo de Pagamento',
        'subtitulo': 'Inscrição',
        'mensagem': (
            'Data: 04/06/2025<br>'
            'Turma 1: 10h às 12h<br>'
            'Turma 2: 14h às 16h<br>'
            'Local: Sala de treinamento da secretaria de ciência e tecnologia, inovação e desenvolvimento econômico - Av. Alberto Braune, 223 - 2º andar.'
        )
    }
    return render(request, 'formfacil/formfacil_form.html', context)

def visualizar_padronizacao_pagamento(request):
    turmas = dict(PadronizacaoPagamentoInscricao.TURMAS)
    registros = {turma: PadronizacaoPagamentoInscricao.objects.filter(turma=key) for key, turma in turmas.items()}
    return render(request, 'formfacil/visualizar_padronizacao_pagamento.html', {'registros': registros, 'turmas': turmas})

def exportar_padronizacao_pagamento_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Padronização do Processo de Pagamento'
    cadastros = PadronizacaoPagamentoInscricao.objects.all().order_by('turma', 'nome')
    ws.append(['Nome', 'Matrícula', 'Secretaria', 'Setor', 'Celular', 'Turma', 'Dt. Inscrição'])
    for item in cadastros:
        ws.append([
            item.nome,
            item.matricula,
            item.secretaria,
            item.setor,
            item.celular,
            item.turma,
            item.dt_inscricao.strftime('%d/%m/%Y %H:%M:%S')
        ])
    col_widths = [25, 15, 20, 20, 15, 15, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    total_rows = ws.max_row
    total_cols = ws.max_column
    first_col = get_column_letter(1)
    last_col = get_column_letter(total_cols)
    table_ref = f"{first_col}1:{last_col}{total_rows}"
    timestamp = int(time.time())
    table_name = f"PadronizacaoPag_{timestamp}"
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=PadronizacaoPagamento.xlsx'
    response.write(output.getvalue())
    return response
