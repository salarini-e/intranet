from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .forms import ItemForm, EditItemForm, AlocaItemForm, RetiraItemForm
from .models import Item, HistoricoItem
from django.db.models import Q
from .decorators import required_almoxarifado
from django.http import HttpResponse
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

@login_required
@required_almoxarifado
def almoxarifado(request):
    query = request.GET.get('filtro', '')
    if query:
        items = Item.objects.filter(
            Q(nome__icontains=query) | Q(id__icontains=query)
        )
    else:
        items = Item.objects.all()
    return render(request, 'almoxarifado.html', { 'items': items, 'filtro': query })

@login_required
@required_almoxarifado
def item_create(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            HistoricoItem.objects.create(
                item=form.instance,
                usuario=request.user.username,
                tipo='ADICAO',
                descricao="Adição de item",
                quantidade_inicial=0,
                quantidade=form.instance.quantidade_total,
                quantidade_final=form.instance.quantidade_total
            )
            return redirect('almoxarifado:almoxarifado')
        else:
            print(form.errors)
    else:
        form = ItemForm()
    return render(request, 'form_item.html', {'form': form})

@login_required
@required_almoxarifado
def item_update(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        form = EditItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('almoxarifado:almoxarifado')
        else:
            print(form.errors)
    else:
        form = EditItemForm(instance=item)
    return render(request, 'form_item_edit.html', {'form': form, "item": item})

@login_required
@required_almoxarifado
def item_delete(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        item.delete()
        return redirect('almoxarifado:almoxarifado')
    return render(request, 'confirm_delete_item.html', {'item': item})

@login_required
@required_almoxarifado
def aloca_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        form = AlocaItemForm(request.POST)
        if form.is_valid():
            alocacao = form.save(commit=False)
            alocacao.item = item
            alocacao.user = request.user.username
            item = item
            quantidade_inicial = item.quantidade_total
            item.quantidade_total += alocacao.quantidade
            item.save()
            alocacao.save()
            # Salva histórico
            HistoricoItem.objects.create(
                item=item,
                usuario=alocacao.user,
                data=alocacao.data,
                tipo='ALOCACAO',
                descricao=alocacao.descricao,
                quantidade_inicial=quantidade_inicial,
                quantidade=alocacao.quantidade,
                quantidade_final=item.quantidade_total
            )
            return redirect('almoxarifado:almoxarifado')
        else:
            print(form.errors)
    else:
        form = AlocaItemForm()
    return render(request, 'aloca_item.html', {'form': form})

@login_required
@required_almoxarifado
def retira_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        form = RetiraItemForm(request.POST)
        if form.is_valid():
            retirada = form.save(commit=False)
            retirada.item = item
            retirada.user = request.user.username
            item = item
            if item.quantidade_total < retirada.quantidade:
                form.add_error('quantidade', 'Quantidade insuficiente em estoque.')
                return render(request, 'form_item.html', {'form': form})
            
            if retirada.quantidade == 0:
                form.add_error('quantidade', 'Quantidade deve ser maior que zero.')
                return render(request, 'form_item.html', {'form': form})
            
            quantidade_inicial = item.quantidade_total
            item.quantidade_total -= retirada.quantidade
            item.save()
            retirada.save()
            # Salva histórico
            HistoricoItem.objects.create(
                item=item,
                usuario=retirada.user,
                data=retirada.data,
                tipo='RETIRADA',
                descricao=retirada.descricao,
                quantidade_inicial=quantidade_inicial,
                quantidade=retirada.quantidade,
                quantidade_final=item.quantidade_total

            )
            return redirect('almoxarifado:almoxarifado')
        else:
            print(form.errors)
    else:
        form = RetiraItemForm()
    return render(request, 'retirada_item.html', {'form': form})

@login_required
@required_almoxarifado
def historico_view(request):
    historicos = HistoricoItem.objects.all().order_by('-data')
    return render(request, 'historico.html', {'historicos': historicos})

@login_required
@required_almoxarifado
def relatorio_xml(request):
    # Cria uma planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Itens"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # azul
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cabeçalhos
    colunas = [
        "ID", "Item", "Usuário", "Data", "Tipo", "Descrição",
        "Qtd. Inicial", "Qtd. Movimentada", "Qtd. Final"
    ]
    ws.append(colunas)

    # Formatar cabeçalho
    for col_num, column_title in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Dados
    historicos = HistoricoItem.objects.all().order_by('-data')

    for row_num, h in enumerate(historicos, start=2):
        valores = [
            h.id,
            h.item.nome,
            h.usuario.username if hasattr(h.usuario, 'username') else h.usuario,
            h.data.strftime('%d/%m/%Y %H:%M'),
            h.get_tipo_display(),
            h.descricao,
            h.quantidade_inicial,
            h.quantidade,
            h.quantidade_final,
        ]
        for col_num, value in enumerate(valores, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = center_alignment
            cell.border = border

    # Ajustar larguras de coluna
    col_widths = [6, 25, 20, 20, 12, 40, 15, 18, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Retorna como arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    data_str = datetime.now().strftime('%d-%m-%Y')
    response['Content-Disposition'] = f'attachment; filename=relatorio_almoxarifado_{data_str}.xlsx'
    wb.save(response)
    return response