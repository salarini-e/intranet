from django.shortcuts import render, redirect
from django.views.decorators.clickjacking import xframe_options_exempt
from datetime import datetime, timedelta
from django.shortcuts import get_object_or_404
from .models import (TipoChamado, Secretaria, Setor, Servidor, Chamado, OSImpressora, OSInternet, OSSistemas, Atendente, Mensagem, OSTelefonia, 
                     PeriodoPreferencial, Pausas_Execucao_do_Chamado, Historico_Designados, chamadoSatisfacao)
from .forms import (CriarChamadoForm, OSInternetForm, OSImpressoraForm, OSSistemasForm, ServidorForm,
                    MensagemForm, AtendenteForm, TipoChamadoForm, OSTelefoniaForm, CriarSetorForm, Form_Agendar_Atendimento,
                    Form_Motivo_Pausa, FormDetalhesDoChamado, FormEditarChamado)
from django.contrib.auth.decorators import login_required
from django.contrib import messages as message
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden, HttpResponseNotFound
from django.utils import timezone
from .functions import enviar_email_atendente, Email_Chamado
from django.urls import reverse
from .functions import obter_filtros, verificar_filtrado, obter_chamados, obter_atendente, obter_opcoes_filtros, paginar_chamados, verificar_chamados_atrasados
from django.core.paginator import Paginator
from autenticacao.functions import clear_tel
import re
from django.db.models import Count, Q
from dateutil.relativedelta import relativedelta
import locale
from django.views.decorators.http import require_POST

from .functions import carregar_novos_filtros, filtrar_chamados
import pandas as pd
import openpyxl
from openpyxl.styles import Font

from .graficos import (date_chamados_por_secretaria, options_chamados_por_secretaria,
                       data_generic, options_generic, date_chamados_por_atendente, 
                       options_chamados_por_atendente, date_chamados_por_mes, options_chamados_por_mes,)



# Define a localidade para português (Brasil)
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

@login_required
def pagina_inicial(request):
    atendentes = Atendente.objects.filter(servidor__user=request.user, ativo=True)
    if atendentes.exists():
        atendente = atendentes.first()
        if atendente.nivel in ['1', '2']:
            return redirect('chamados:tickets')
        elif atendente.nivel=='0':
            return redirect('chamados:criar_chamado_escolher')            
        return redirect('chamados:tickets')
    return redirect('chamados:criar_chamado_escolher')

@login_required
def ordenarPorNaoFechados(request):
    if 'ordenacao' in request.session:
        if request.session['ordenacao'] == True:
            request.session['ordenacao'] = False
        else:
            request.session['ordenacao'] = True
    return redirect('chamados:tickets')

@login_required
def criarChamadoEscolher(request):    
    context = {
               'tipos': TipoChamado.objects.all(),
               }
    return render(request, 'chamados/criar-chamado-escolher.html', context)

@login_required
def index(request):
    filtros = obter_filtros(request)
    filtrado = verificar_filtrado(request, filtros)
    chamados = obter_chamados(request, filtros)
    atendente = obter_atendente(request)
    paginacao = paginar_chamados(request, chamados)
    chamados_atrasados_data_agendada, chamados_atrasados_trinta_dias = verificar_chamados_atrasados()

    kpi = {
        'total': chamados.count(),
        'abertos': chamados.filter(status='1').count(),
        'pendentes': chamados.filter(status='2').count(),
        'fechados_finalizados': chamados.filter(status__in=['3', '4']).count()
    }
    
    context = {
        'tipos': TipoChamado.objects.all(),
        'filtrado': filtrado,
        'filtros': obter_opcoes_filtros(),
        'chamados': paginacao,
        'atendente': atendente,        
        'chamados_atrasados': {
                                'exists': any([chamados_atrasados_data_agendada, chamados_atrasados_data_agendada]),
                                'trintadias': chamados_atrasados_trinta_dias, 
                                'agendado': chamados_atrasados_data_agendada
                                },
        'kpi':  kpi
    }
    return render(request, 'chamados/index.html', context)

def zerar_filtros(request):
    if 'filtros' in request.session:
        del request.session['filtros']
    return redirect('chamados:index')  # Redireciona de volta para a página inicial onde os filtros são aplicados

@login_required
def criarChamado(request, sigla):
    atendentes = Atendente.objects.filter(servidor__user=request.user, ativo=True)
    if atendentes.exists():
        atendente = atendentes.first()
        if atendente.nivel == '1':
            return redirect('chamados:tickets')

    def itel(v):
        # Remove todos os caracteres que não são dígitos
        v = re.sub(r'\D', '', v)

        # Adiciona parênteses e espaço para o DDD
        v = re.sub(r'^(\d{2})(\d)', r'(\1) \2', v)

        # Verifica o comprimento da string para formatar o número corretamente
        if len(v) == 14:  # Formato DDD + 9 dígitos (celular)
            v = re.sub(r'(\d{5})(\d)', r'\1-\2', v)  # Coloca hífen entre o 5º e 6º dígitos
        else:  # Formato DDD + 8 dígitos (fixo)
            v = re.sub(r'(\d{4})(\d)', r'\1-\2', v)  # Coloca hífen entre o 4º e 5º dígitos

        return v
    # try:
    #     servidor = Servidor.objects.get(user=request.user)
    # except Servidor.DoesNotExist:
    #     servidor = None
    is_atendente = Atendente.objects.filter(servidor__user=request.user, ativo = True).exists() 

    if request.user.is_superuser or is_atendente:
        template_name = 'chamados/chamado-criar-ti.html'
        servidor = Servidor.objects.get(user=request.user)
        tipo = TipoChamado.objects.get(sigla=sigla)
        initial_data = {
            'tipo': tipo,
            'user_inclusao': servidor.id,
            'setor_id': '-',
            'criado_pelo_servidor': '0',
        }
    else:
        template_name = 'chamados/chamado-criar-servidor.html'
        servidor = Servidor.objects.get(user=request.user)
        tipo = TipoChamado.objects.get(sigla=sigla)
        telefone_formatado = itel(servidor.telefone)  # Formata o telefone
        initial_data = {
            'secretaria': servidor.setor.secretaria.id,
            'setor': '-',
            'telefone': telefone_formatado,  # Usa o telefone formatado
            'tipo': tipo,
            'requisitante': servidor.id, 
            'user_inclusao': servidor.id,
            'criado_pelo_servidor': '1',
        }

    if request.POST:
        print(request.POST)
        form = CriarChamadoForm(request.POST, request.FILES, user=request.user)
        if sigla == 'TEL':
            form_ext = OSTelefoniaForm(request.POST)
        if form.is_valid() and (sigla != 'TEL' or form_ext.is_valid()):
            chamado = form.save(commit=False)
            chamado.user_inclusao = servidor

            chamado.telefone = clear_tel(form.cleaned_data['telefone'])
            chamado.profissional_designado = form.cleaned_data.get('profissional_designado')
            
            chamado.save()
            chamado.gerar_hash()
            chamado.gerar_protocolo()
            if sigla == 'TEL':
                try:
                    OSTelefonia.objects.create(chamado=chamado, ramal=request.POST['ramal'])
                except Exception as e:
                    print(e)


            message.success(request, 'Chamado {} criado com sucesso! Não esqueça de responder nossa pesquisa de satisfação quando seu chamado for finalizado!'.format(chamado.n_protocolo))
            mensagem, status = Email_Chamado(chamado).chamado_criado()
            if status == 400:
                message.error(request, mensagem)

            try:
                chamado.gerar_notificacoes()
            except ValueError as e:
                message.error(request, str(e))
            return redirect('chamados:tickets')        
    else:
        form = CriarChamadoForm(initial=initial_data, user=request.user)

    context = {
        'tipos': TipoChamado.objects.all(),
        'form': form,
        'form_setor': CriarSetorForm(prefix='setor', initial={'user_inclusao': request.user.id}),
        'form_user': ServidorForm(prefix='servidor', initial={'user_inclusao': request.user.id})
    }
    if sigla == 'TEL' and request.method != 'POST':
        context['form_ext'] = OSTelefoniaForm()
    elif sigla == 'TEL' and request.method == 'POST':
        context['form_ext'] = form_ext
    return render(request, template_name, context)

@login_required
def detalhes(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    servidor = Servidor.objects.get(user=request.user)
    tipos_chamados = TipoChamado.objects.all()

    if request.method == 'POST':        
        form = MensagemForm(request.POST, request.FILES)
        if form.is_valid():
            mensagem = form.save(commit=False)
            mensagem.chamado = chamado
            mensagem.user_inclusao = servidor
            mensagem.confidencial = request.POST.get('confidencial') == '1'
            mensagem.save()            
            mensagem.notificar_nova_mensagem()
            chamado.dt_atualizacao = timezone.now() 
            chamado.save() 
            # try:
            #     mensagem, status = Email_Chamado(chamado).mensagem_criada(mensagem)
            # except Exception as e:
            #     print(e)
            # if status == 400:
            #     message.error(request, mensagem)

            message.success(request, 'Mensagem enviada com sucesso!')
        # else:
        #     print(form.errors)

    # extensoes = {
    #     'IMP': OSImpressora,
    #     'INT': OSInternet,
    #     'SIS': OSSistemas,
    #     'TEL': OSTelefonia,
    # }

    # Função para formatar o telefone
    def itel(v):
        v = re.sub(r'\D', '', v)
        v = re.sub(r'^(\d{2})(\d)', r'(\1) \2', v)
        if len(v) == 14:
            v = re.sub(r'(\d{5})(\d)', r'\1-\2', v)
        else:
            v = re.sub(r'(\d{4})(\d)', r'\1-\2', v)
        return v
    
    requisitante_telefone = itel(chamado.requisitante.telefone)
    
    context = {
        'chamado': chamado,    
        'servidor': servidor,    
        'atendentes': Atendente.objects.filter(ativo=True),
        'atendente': Atendente.objects.filter(servidor=servidor),   
        'mensagens': Mensagem.objects.filter(chamado=chamado),   
        'prioridades': chamado.PRIORIDADE_CHOICES,
        'status': chamado.STATUS_CHOICES,     
        'is_atendente': Atendente.objects.filter(servidor = servidor, ativo = True).exists(),
        'helpdesk_or_adm': Atendente.objects.filter(servidor = servidor, nivel__in = ['0', '2'],ativo = True).exists(),
        'tipos': tipos_chamados,
        'tipos_chamados': tipos_chamados,
        'requisitante_telefone': requisitante_telefone,
        'form_relatorio': FormDetalhesDoChamado(instance=chamado)
    }
    if chamado.tipo.sigla == 'TEL':
        context['ramais'] = OSTelefonia.objects.get(chamado=chamado).ramal
    return render(request, 'chamados/detalhes.html', context)

def editar_chamado(request, hash_chamado):
    servidor = Servidor.objects.get(user=request.user)
    if not Atendente.objects.filter(servidor = servidor, nivel__in = ['0', '2'],ativo = True).exists():
        return HttpResponseForbidden('Você não tem permissão para acessar esta página!')
    chamado = get_object_or_404(Chamado, hash=hash_chamado)
    if request.method == 'POST':
        form = FormEditarChamado(request.POST, instance=chamado)
        if form.is_valid():
            chamado = form.save()
            return redirect('chamados:detalhes', hash=chamado.hash)
        else:
            print(form.errors)
    else:
        form = FormEditarChamado(instance=chamado)
    context = {
        'form': form,
        'chamado': chamado,
    }
    return render(request, 'chamados/editar_chamado.html', context)
        

from django.db.models import Count, Q, Avg, Max, Min, F

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, F
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Chamado, Secretaria, Atendente, TipoChamado

@login_required
def gerar_relatorio(request):
    # Filtra chamados de acordo com os critérios do request
    chamados = filtrar_chamados(request)
    protocolos = [chamado.n_protocolo for chamado in chamados]
    chamados_queryset = Chamado.objects.filter(n_protocolo__in=protocolos)

    # Dados estáticos ou de referência
    secretarias = Secretaria.objects.all()
    atendentes = Atendente.objects.filter(ativo=True)
    status_choices = Chamado.STATUS_CHOICES
    prioridade_choices = Chamado.PRIORIDADE_CHOICES
    tipos_chamados = TipoChamado.objects.all()

    # Estatísticas gerais
    total_chamados = Chamado.objects.count()

    # Quantidade de chamados por status
    chamados_por_status = {
        status: chamados_queryset.filter(status=str(status)).count()
        for status in [0, 2, 4, 5]  # Define os status relevantes
    }

    # Chamados agrupados por tipo
    chamados_por_tipo = (
        chamados_queryset.values('tipo__nome')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    # print(chamados_por_tipo)

    # Chamados agrupados por prioridade
    chamados_por_prioridade = (
        chamados_queryset.values('prioridade')
        .annotate(total=Count('id'))
    )

    # Tempo médio de resolução para chamados finalizados
    tempo_medio_resolucao = (
        chamados_queryset.filter(
            status='4',  # Finalizado
            dt_inicio_execucao__isnull=False,
            dt_execucao__isnull=False
        )
        .annotate(
            tempo_resolucao=F('dt_execucao') - F('dt_inicio_execucao')
        )
        .aggregate(tempo_medio=Avg('tempo_resolucao'))['tempo_medio']
    )

    # Atendentes com mais chamados atribuídos
    atendentes_ativos = (
        chamados_queryset.values('profissional_designado__nome_servidor')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Configuração de paginação
    page = request.GET.get('page', 1)
    paginator = Paginator(chamados, 10)
    page_obj = paginator.get_page(page)

    # Remove parâmetros desnecessários da URL
    current_query = request.GET.copy()
    current_query.pop('page', None)

    # Contexto para o template
    context = {
        'tipos': tipos_chamados,
        'chamados': page_obj,
        'secretarias': secretarias.order_by('apelido'),
        'atendentes': atendentes,
        'status_choices': status_choices,
        'prioridade_choices': prioridade_choices,
        'current_query': current_query,
        'is_atendente': Atendente.objects.filter(
            servidor__user=request.user, ativo=True
        ).exists(),

        # Dados estatísticos
        'total_chamados': total_chamados,
        'chamados_por_status': chamados_por_status,
        'chamados_por_tipo': chamados_por_tipo,
        'chamados_por_prioridade': chamados_por_prioridade,
        'tempo_medio_resolucao': tempo_medio_resolucao,
        'atendentes_ativos': atendentes_ativos,
    }

    return render(request, 'chamados/relatorio.html', context)

@login_required
def api_relatorio(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    if request.method == 'POST':
        form = FormDetalhesDoChamado(request.POST, instance=chamado)
        # print(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'status': 200, 'message': 'Relatório salvo com sucesso!'})
        else:
            return JsonResponse({'status': 400, 'message': 'Erro ao salvar relatório!'})
    return JsonResponse({'status': 400, 'message': 'Método não permitido!'})
from django.template.loader import render_to_string

@login_required
def api_montarFormRelatorio(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    form = FormDetalhesDoChamado(instance=chamado)
    context = {
        'form_relatorio': form,        
    }
    # Renderiza o template como string
    html = render_to_string('chamados/form_relatorio.html', context, request=request)
    return JsonResponse({'status': 200, 'html': html})
@login_required
def detalhes_imprimir(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    servidor = Servidor.objects.get(user=request.user)
    if request.method == 'POST':
        
        form = MensagemForm(request.POST, request.FILES)
        if form.is_valid():
            mensagem = form.save(commit=False)
            mensagem.chamado = chamado
            mensagem.user_inclusao = servidor
            mensagem.save()            
            message.success(request, 'Mensagem enviada com sucesso!')
        # else:
        #     print(form.errors)

    # extensoes = {
    #     'IMP': OSImpressora,
    #     'INT': OSInternet,
    #     'SIS': OSSistemas,
    #     'TEL': OSTelefonia,
    # }
    # if chamado.tipo.sigla in extensoes:
    #     extensao = extensoes[chamado.tipo.sigla].objects.get(chamado=chamado)
    # else:
    #     extensao = None

    context = {
        'chamado': chamado,
        # 'ext': extensao,    
        'servidor': servidor,    
        'atendentes': Atendente.objects.filter(ativo=True),
        'atendente': Atendente.objects.filter(servidor=servidor),   
        'mensagens': Mensagem.objects.filter(chamado=chamado),   
        'prioridades': chamado.PRIORIDADE_CHOICES,
        'status': chamado.STATUS_CHOICES,   
              
    }
    return render(request, 'chamados/detalhes-imprimir.html', context)

@login_required
def attChamado(request, hash):
    servidor = Servidor.objects.get(user=request.user)
    atendente = Atendente.objects.filter(servidor=servidor)
    if atendente.exists():
        chamado = Chamado.objects.get(hash=hash)
        if request.POST:
            atributo = request.POST['atributo']
            if atributo == 'status':
                chamado.status = request.POST['valor']   
                if request.POST['valor'] == '4':
                    chamado.dt_fechamento = timezone.now()             
            elif atributo == 'prioridade':
                chamado.prioridade = request.POST['valor']
            elif atributo == 'atendente':
                chamado.profissional_designado = Atendente.objects.get(id=request.POST['valor'])                
                
                mensagem, status = Email_Chamado(chamado).notificar_solicitante('Seu chamado foi atualizado!')
                if status == 400:
                    message.error(request, mensagem)
                else:
                    message.success(request, mensagem)

                mensagem_atendente, status_atendente = Email_Chamado(chamado).atribuir_atendente('Chamado atribuído a você!')
                if status_atendente == 400:
                    message.error(request, mensagem_atendente)
                    
                elif status_atendente == 200:
                    message.success(request, mensagem_atendente)
                
                
            else:                
                return JsonResponse({'status': 400})
            chamado.save()  
            message.success(request, f'{atributo.capitalize()} atualizado com sucesso!')
            return JsonResponse({'status': 200})    
    return JsonResponse({'status': 403})

@login_required
def iniciar_atendimento(request, hash):    
    chamado = Chamado.objects.get(hash=hash)
    pagina_anterior = request.META.get('HTTP_REFERER', '/')
    if not chamado.dt_agendamento:
        chamado.dt_agendamento =  timezone.now()     
    chamado.dt_inicio_execucao = timezone.now()
    chamado.status = '1'
    chamado.save()
    message.success(request, f'Atendimento iniciado ao chamado {chamado.n_protocolo}!')
    mensagem, status = Email_Chamado(chamado).notificar_solicitante('Seu chamado está em atendimento!')
    if status == 400:
        message.error(request, mensagem)
            
    return redirect(pagina_anterior)

@login_required
def retomar_atendimento(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    intervalo = Pausas_Execucao_do_Chamado.objects.filter(chamado=chamado).last()
    intervalo.dt_fim = timezone.now()
    intervalo.user_fim = Servidor.objects.get(user=request.user)
    intervalo.save()
    chamado.status = '1'
    chamado.save()
    message.success(request, f'Retomado o atendimento do chamado {chamado.n_protocolo}!')
    mensagem, status = Email_Chamado(chamado).notificar_solicitante('Seu chamado está em atendimento!')
    if status == 400:
        message.error(request, mensagem)
        
    return redirect('chamados:index')


@login_required
def pausar_atendimento(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    intervalo = Pausas_Execucao_do_Chamado.objects.create(
        chamado=chamado,
        dt_inicio=timezone.now(),        
        user_inclusao=Servidor.objects.get(user=request.user)
    
    )    
    chamado.status = '2'
    chamado.save()
    message.success(request, f'Atendimento do chamado {chamado.n_protocolo} pausado!')
    mensagem, status = Email_Chamado(chamado).notificar_solicitante('O atendimento do seu chamado foi interrompido!')
    if status == 400:
        message.error(request, mensagem)
    return redirect('chamados:motivo', hash=hash)

@login_required
def declarar_motivo_pausa(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    instance = Pausas_Execucao_do_Chamado.objects.filter(chamado=chamado, user_inclusao=Servidor.objects.get(user=request.user)).last()
    if request.method == 'POST':
        form = Form_Motivo_Pausa(request.POST, instance=instance)
        if form.is_valid():
            form.save()            
            return redirect('chamados:index')
    else:
        form = Form_Motivo_Pausa()
    context = {
        'form': form, 
               'titulo': f'Motivo da paus do chamado {chamado.n_protocolo}', 
               'back_url': reverse('chamados:index'),
               'submit_text': 'Enviar'
               }
    return render(request, 'chamados/generic_form.html', context)

@login_required
def finalizar_atendimento(request, hash):
    chamado = Chamado.objects.get(hash=hash)
    chamado.dt_execucao = timezone.now()
    chamado.status = '4'
    chamado.save()
    message.success(request, f'Atendimento finalizado ao chamado {chamado.n_protocolo}!')
    mensagem, status = Email_Chamado(chamado).notificar_solicitante('Seu chamado foi finalizado!')
    if status == 400:
        message.error(request, mensagem)
    return redirect('chamados:index')

@login_required
def criar_periodos(request):
    if PeriodoPreferencial.objects.all().exists():
        return HttpResponse("OK. Periodos já criados.")

    PeriodoPreferencial.objects.create(nome="Manhã")
    PeriodoPreferencial.objects.create(nome="Tarde")    
    return HttpResponse("OK. Periodos criados com sucesso!")

@login_required
def api_criar_setor(request):
    if request.method == 'POST':
        # print(request.POST)
        data = request.POST
        setor = Setor.objects.create(
            nome=data['nome'],
            apelido=data['apelido'],
            sigla=data['sigla'],
            cep=data['cep'],
            bairro=data['bairro'],
            endereco=data['endereco'],
            secretaria=Secretaria.objects.get(id=data['secretaria']),
        )
        return JsonResponse({'status': 200, 'message': 'Setor criado com sucesso!'})
    return JsonResponse({'status': 400, 'message': 'Erro ao criar setor!'})

@login_required
def api_mudar_status(request):
    servidor = Servidor.objects.filter(user=request.user).last()
    if request.user.is_superuser or Atendente.objects.filter(servidor=servidor).exists():
        pass
    else:
        return JsonResponse({'status': 403, 'message': 'Acesso negado!'})
    if request.method == 'POST':
        data = request.POST
        try:
            chamado = Chamado.objects.get(hash=data['hash'])
            chamado.status = data['status']
            chamado.dt_atualizacao = timezone.now()
            chamado.save()
            chamado.notificar_status_alterado() 
            # print(chamado.status)
            if chamado.status == '4':                                
                mensagem, status = Email_Chamado(chamado).chamado_finalizado()
                # Email_Chamado(chamado).chamado_finalizado()
                if status == 400:
                    message.error(request, mensagem)
            return JsonResponse({'status': 200, 'message': 'Status atualizado com sucesso!', 'display_status': chamado.get_status_display(), 'id': chamado.id})
        except Exception as E:
            print(E)
            return JsonResponse({'status': 400, 'message': 'Erro ao atualizar status!'})
    return JsonResponse({'status': 400, 'message': 'Método não permitido!'})

@login_required
def api_mudar_prioridade(request):
    servidor = Servidor.objects.filter(user=request.user).last()
    if request.user.is_superuser or Atendente.objects.filter(servidor=servidor).exists():
        pass
    else:
        return JsonResponse({'status': 403, 'message': 'Acesso negado!'})
    if request.method == 'POST':
        data = request.POST
        try:            
            chamado = Chamado.objects.get(hash=data['hash'])
            chamado.prioridade = data['prioridade']
            chamado.dt_atualizacao = timezone.now()
            chamado.save()
            return JsonResponse({'status': 200, 'message': 'Prioridade atualizada com sucesso!', 'display_prioridade': chamado.get_prioridade_display(), 'id': chamado.id})
        except Exception as E:
            print(E)
            return JsonResponse({'status': 400, 'message': 'Erro ao atualizar prioridade!'})
    return JsonResponse({'status': 400, 'message': 'Método não permitido!'})

@login_required
def api_mudar_atendente(request):
    servidor = Servidor.objects.filter(user=request.user).last()
    if request.user.is_superuser or Atendente.objects.filter(servidor=servidor).exists():
        pass
    else:
        return JsonResponse({'status': 403, 'message': 'Acesso negado!'})
    
    if request.method == 'POST':
        data = request.POST
        try:            
            chamado = Chamado.objects.get(hash=data['hash'])
            
            # Buscar a instância do atendente
            atendente = Atendente.objects.get(id=data['atendente'])
            
            # Atribuir a instância do atendente ao chamado
            chamado.profissional_designado = atendente
            chamado.dt_atualizacao = timezone.now()
            chamado.save()
            chamado.notificar_profissional_designado_alterado()
            Historico_Designados.objects.create(
                chamado=chamado,
                atendente=atendente,
                user_inclusao=request.user
            )
            # print("Chamado id", chamado.id)
            return JsonResponse({
                'status': 200,
                'message': 'Atendente atualizado com sucesso!',
                'display_atendente':  chamado.profissional_designado.nome_servidor,
                'id': chamado.id
            })
        except Exception as E:
            print(E)
            return JsonResponse({'status': 400, 'message': 'Erro ao atualizar atendente!'})
    
    return JsonResponse({'status': 400, 'message': 'Método não permitido!'})

@login_required
def api_criar_servidor(request):
    if request.method == 'POST':
        # print(request.POST)
        form = ServidorForm(request.POST)                    
        if form.is_valid():
            servidor = form.save()    
            servidor.user = form.create_user()
            servidor.user_inclusao = request.user
            servidor.save()
        else:
            # print(form.errors)
            return JsonResponse({'status': 400, 'message': 'Erro ao criar usuário!'})
        
        return JsonResponse({'status': 200, 'message': 'Usuário criado com sucesso!'})        
    return JsonResponse({'status': 400})

@login_required
def agendar_atendimento(request, hash):
    instance = Chamado.objects.get(hash=hash)
    if request.method == 'POST':
        form = Form_Agendar_Atendimento(request.POST, instance=instance)
        if form.is_valid():
            chamado=form.save()
            message.success(request, 'Atendimento agendado com sucesso!')
            mensagem, status = Email_Chamado(chamado).notificar_solicitante('O atendimento do seu chamado foi agendado!')
            if status == 400:
                message.error(request, mensagem)
            return redirect('chamados:detalhes', hash=hash)
    else:
        form = Form_Agendar_Atendimento()
    context = {'form': form, 
               'titulo': f'Chamado {instance.n_protocolo}', 
               'back_url': reverse('chamados:detalhes', args=[instance.hash]),
               'submit_text': 'Agendar'
               }
    return render(request, 'chamados/generic_form.html', context)



@login_required
def tickets(request):
    if not 'ordenacao' in request.session:
        request.session['ordenacao'] = False
    
    if request.method == 'POST':
        carregar_novos_filtros(request)

    chamados = filtrar_chamados(request)

    # CONDIÇÃO PARA O PAINEL DE CONTROLE (LISTAR TODOS OS TICKETS QUE NÃO ESTÃO ATRIBUÍDOS)
    if request.GET.get('profissional_designado__isnull') == 'True':
        # Filtra apenas os chamados onde o profissional designado é None
        chamados = [chamado for chamado in chamados if chamado.profissional_designado is None]

    # Filtro para os últimos 30 dias
    if request.GET.get('ultimos_30_dias') == 'True':
        data_limite = timezone.now().replace(tzinfo=None) - timedelta(days=30)
        chamados = [chamado for chamado in chamados if chamado.dt_inclusao >= data_limite]

    if request.GET.get('nao_resolvidos') == 'True':
        chamados = [chamado for chamado in chamados if chamado.status != '4']

    # PARTE PARA LISTAR OS TICKETS NA ABA DE TICKETS NÃO RESOLVIDOS DO PAINEL DE CONTROLE
    status = request.GET.get('status')
    tipo = request.GET.get('tipo')
    profissional_designado = request.GET.get('profissional_designado')
    # CONDIÇÃO PARA O PAINEL DE CONTROLE (LISTAR OS TICKETS DE ACORDO COM O TIPO, STATUS, PROFISSIONAL DESIGINADO) -> TICKETS NÃO RESOLVIDOS
    if tipo:
        if status:
            chamados = [chamado for chamado in chamados if chamado.tipo_id == int(tipo) and chamado.status == str(status)]
        else:
            chamados = [chamado for chamado in chamados if chamado.tipo_id == int(tipo)]
    
    elif profissional_designado:
        if status: 
            chamados = [chamado for chamado in chamados if chamado.profissional_designado_id == int(profissional_designado) and chamado.status == str(status)]
        else: 
            chamados = [chamado for chamado in chamados if chamado.profissional_designado_id == int(profissional_designado)]
    elif status and not tipo and not profissional_designado:
        chamados = [chamado for chamado in chamados if chamado.status == str(status)]
        

    secretarias = Secretaria.objects.all()
    atendentes = Atendente.objects.all()
    tipos_chamados = TipoChamado.objects.all()

    paginator = Paginator(chamados, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # MANTER O FILTRO NA URL MESMO AO TROCAR A PÁGINA DO PAGINATOR
    current_query = request.GET.copy()
    if 'page' in current_query:
        del current_query['page']  # Remover o parâmetro 'page'

    status_choices = Chamado.STATUS_CHOICES
    prioridade_choices = Chamado.PRIORIDADE_CHOICES
    context = {
        'tipos': tipos_chamados,
        'chamados': page_obj,
        'secretarias': secretarias.order_by('apelido'),
        'atendentes': atendentes,
        'tipos_chamados': tipos_chamados,
        'status_choices': status_choices,
        'prioridade_choices': prioridade_choices,
        'current_query': current_query.urlencode(),
        'is_atendente': Atendente.objects.filter(servidor__user=request.user, ativo=True).exists()
    }

    return render(request, 'chamados/tickets.html', context)


def dados_graficos(chamados, data_inicial, data_final, tempo):
    labels=[]
    dados_abertos =[]
    dados_fechados = []
    while data_inicial<= data_final:
        if tempo == 'Hoje':
            label = data_inicial.strftime('%d/%m/%y')
            proximo_periodo = timedelta(days=1)
        elif tempo == 'Uma-semana':
            label = data_inicial.strftime('%d/%m/%y')
            proximo_periodo = timedelta(days=1)
        elif tempo == 'Um-mes':
            label = data_inicial.strftime('%d/%m/%y')
            proximo_periodo = timedelta(days=1)
        elif tempo =='Este-ano':
            label = data_inicial.strftime('%b')
            proximo_periodo = timedelta(days=31)
        elif tempo=='Um-ano':
            label = data_inicial.strftime('%b/%y')
            proximo_periodo = timedelta(days=30)

        if tempo =='Hoje':
            for hour in range(0, 24, 2):
                label= f'{hour:02d}:00'
                labels.append(label)
                
                inicio_hora = data_inicial + timedelta(hours=hour)
                fim_hora = inicio_hora + timedelta(hours=2)
                
                chamados_abertos_intervalo = chamados.filter(dt_inclusao__gte=inicio_hora, dt_inclusao__lt=fim_hora, status='0').count()
                chamados_fechados_intervalo = chamados.filter(dt_inclusao__gte=inicio_hora, dt_inclusao__lt=fim_hora, status='3').count()
                
                dados_abertos.append(chamados_abertos_intervalo)
                dados_fechados.append(chamados_fechados_intervalo)

            dados_abertos.insert(0, chamados_abertos_intervalo) 
            dados_fechados.insert(0, chamados_fechados_intervalo)
        else:
            chamados_periodo = chamados.filter(dt_inclusao__gte=data_inicial, dt_inclusao__lt=data_inicial + proximo_periodo)
            chamados_abertos = chamados_periodo.filter(status='0').count()
            chamados_fechados = chamados_periodo.filter(status='3').count()

            labels.append(label)
            dados_abertos.append(chamados_abertos)
            dados_fechados.append(chamados_fechados)
        data_inicial += proximo_periodo

    return labels, dados_abertos, dados_fechados

def dados_graficos_tipo(chamados, data_inicial, data_final, tempo, tipo):
    labels=[]
    dados_abertos =[]
    dados_fechados = []
    while data_inicial<= data_final:
        if tempo == 'Hoje':
            label = data_inicial.strftime('%d/%m/%y')
            proximo_periodo = timedelta(days=1)
        elif tempo == 'Uma-semana':
            label = data_inicial.strftime('%d/%m/%y')
            proximo_periodo = timedelta(days=1)
        elif tempo == 'Um-mes':
            label = data_inicial.strftime('%d/%m/%y')
            proximo_periodo = timedelta(days=1)
        elif tempo =='Este-ano':
            label = data_inicial.strftime('%b')
            proximo_periodo = timedelta(days=31)
        elif tempo=='Um-ano':
            label = data_inicial.strftime('%b/%y')
            proximo_periodo = timedelta(days=30)

        if tempo =='Hoje':
            for hour in range(0, 24, 2):
                label= f'{hour:02d}:00'
                labels.append(label)
                
                inicio_hora = data_inicial + timedelta(hours=hour)
                fim_hora = inicio_hora + timedelta(hours=2)
                
                chamados_abertos_intervalo = chamados.filter(dt_inclusao__gte=inicio_hora, dt_inclusao__lt=fim_hora, status='0', tipo=tipo).count()
                chamados_fechados_intervalo = chamados.filter(dt_inclusao__gte=inicio_hora, dt_inclusao__lt=fim_hora, status='3', tipo=tipo).count()
                
                dados_abertos.append(chamados_abertos_intervalo)
                dados_fechados.append(chamados_fechados_intervalo)

            dados_abertos.insert(0, chamados_abertos_intervalo) 
            dados_fechados.insert(0, chamados_fechados_intervalo)
        else:
            chamados_periodo = chamados.filter(dt_inclusao__gte=data_inicial, dt_inclusao__lt=data_inicial + proximo_periodo, tipo=tipo)
            chamados_abertos = chamados_periodo.filter(status='0').count()
            chamados_fechados = chamados_periodo.filter(status='3').count()

            labels.append(label)
            dados_abertos.append(chamados_abertos)
            dados_fechados.append(chamados_fechados)
        
        data_inicial += proximo_periodo

    return labels, dados_abertos, dados_fechados

@xframe_options_exempt
@login_required
def painel_controle(request):
    chamados = Chamado.objects.all()
    total_chamados = chamados.count()
    total_chamados_abertos = chamados.filter(status='0').count()
    chamados_abertos_30dias = chamados.filter(dt_inclusao__gte=datetime.now().replace(tzinfo=None) - timedelta(days=30), status='0').count()
    chamados_fechados_30dias = chamados.filter(dt_execucao__gte=datetime.now().replace(tzinfo=None) - timedelta(days=30)).count()
    # chamados_finalizados_30dias =  chamados.filter(dt_execucao__gte=datetime.now().replace(tzinfo=None) - timedelta(days=30), status = '4').count()
    chamados_finalizados_30dias =  chamados.filter(dt_atualizacao__gte=datetime.now().replace(tzinfo=None) - timedelta(days=30), status = '4').count()
    media_diaria = total_chamados / 30
    data_atual = datetime.now()
    tipos_chamados = TipoChamado.objects.all()
    tres_meses_atras = data_atual - timedelta(days=90)
    count_abertos = chamados.filter(status='0').count()
    count_em_atendimento = chamados.filter(status='1').count()
    count_pendentes = chamados.filter(status='2').count()
    count_fechados = chamados.filter(status='3').count()
    count_finalizados = chamados.filter(status='4').count()
    total_nao_resolvidos = total_chamados - count_finalizados - chamados.filter(status='5').count() - chamados.filter(status='6').count()
    total_nao_atribuidos = chamados.filter(profissional_designado__isnull=True).count()

    # Calcular a porcentagem de não atribuídos
    porcentagem_nao_atribuidos = (total_nao_atribuidos / total_chamados * 100) if total_chamados > 0 else 0
    # Calcular a porcentagem de não resolvidos
    porcentagem_nao_resolvidos = (total_nao_resolvidos / total_chamados * 100) if total_chamados > 0 else 0
    #Calcular a porcentagem de abertos nos ultimos trinta dias em relação ao total
    # porcentagem_abertos_ultimos_trinta_dias =   (chamados_abertos_30dias / total_chamados * 100) if total_chamados > 0 else 0
    # Preparando dados para o gráfico de barras
    chamados_por_tipo = [{'tipo': tipo.nome, 'quantidade': chamados.filter(tipo=tipo, status='0').count()} for tipo in tipos_chamados]
    # Criando dicionário para armazenar quantidades de chamados abertos por tipo
    chamados_abertos_por_tipo = {
        tipo.nome: chamados.filter(tipo=tipo, status='0').count() for tipo in tipos_chamados
    }
    secretarias = Secretaria.objects.all()
    chamados_abertos_por_secretaria = {}

    for chamado in chamados:
        secretaria = chamado.secretaria
        if secretaria is not None and secretaria != "":
            chamados_abertos_por_secretaria[secretaria.sigla] = chamados.filter(secretaria=secretaria, status='0').count()
        else:
            if 'Sem secretaria definida' not in chamados_abertos_por_secretaria:
                chamados_abertos_por_secretaria['Sem secretaria definida'] = 0
            chamados_abertos_por_secretaria['Sem secretaria definida'] += 1


    # Para incluir o total de chamados sem secretaria, faça uma consulta separada se necessário
    chamados_abertos_por_secretaria['Sem secretaria definida'] = chamados.filter(secretaria=None, status='0').count()
    
    # Chamados criados entre 30 e 60 dias atrás
    chamados_abertos_30_a_60_dias = chamados.filter(
        dt_inclusao__lt=datetime.now().replace(tzinfo=None) - timedelta(days=30),
        dt_inclusao__gte=datetime.now().replace(tzinfo=None) - timedelta(days=60), status='0'
    ).count()

    # Verifica se houve aumento ou diminuição
    aumento_chamados_abertos_30_dias= chamados_abertos_30dias > chamados_abertos_30_a_60_dias

    # Cálculo da porcentagem em relação ao total de chamados
    porcentagem_abertos_30dias = (chamados_abertos_30dias / total_chamados * 100) if total_chamados > 0 else 0
    porcentagem_abertos_30_a_60_dias = (chamados_abertos_30_a_60_dias / total_chamados * 100) if total_chamados > 0 else 0

    variacao_30_a_60_dias = abs(porcentagem_abertos_30dias - porcentagem_abertos_30_a_60_dias)
    

    # DADOS PARA OS GRÁFICOS DE TODOS OS CHAMADOS
    data_atual = datetime.now()
    um_mes_atras = data_atual - timedelta(days=30)
    uma_semana_atras = data_atual - timedelta(days=7)
    um_ano = data_atual - timedelta(days=365)
    inicio_do_ano = datetime(data_atual.year, 1, 1)
    esse_ano = inicio_do_ano

    labels_atendimentos_realizados = []
    dados_abertos_atendimentos_realizados = []
    esse_ano_atendimentos = esse_ano
    while esse_ano_atendimentos <= data_atual:
        # Define o primeiro dia do mês e o primeiro dia do mês seguinte
        proximo_mes = esse_ano_atendimentos + relativedelta(months=1)
        
        # Filtra os chamados do mês atual e com status diferente de 0
        chamados_esse_ano = chamados.filter(
            dt_inclusao__gte=esse_ano_atendimentos,
            dt_inclusao__lt=proximo_mes
        )
        
        # Conta os chamados com status diferente de '0'
        chamados_atendidos_esse_ano = chamados_esse_ano.filter(~Q(status='0')).count()

        # Só adiciona o mês se houver chamados atendidos
        if chamados_atendidos_esse_ano > 0:
            label_atendimento_realizado = esse_ano_atendimentos.strftime('%b')
            labels_atendimentos_realizados.append(label_atendimento_realizado)
            dados_abertos_atendimentos_realizados.append(chamados_atendidos_esse_ano)
        else:
            label_atendimento_realizado = esse_ano_atendimentos.strftime('%b')
            labels_atendimentos_realizados.append(label_atendimento_realizado)
            dados_abertos_atendimentos_realizados.append(0)

        # Avança para o próximo mês
        esse_ano_atendimentos = proximo_mes
    
    semanas_ano, dados_abertos_um_ano, dados_fechados_um_ano = dados_graficos(chamados, um_ano, data_atual,'Um-ano' )
    meses_esse_ano, dados_abertos_esse_ano, dados_fechados_esse_ano = dados_graficos(chamados,esse_ano, data_atual, "Este-ano" )
    semanas_mes, dados_abertos_um_mes,dados_fechados_um_mes = dados_graficos(chamados, um_mes_atras, data_atual, "Um-mes")
    uma_semana, dados_abertos_uma_semana, dados_fechados_uma_semana = dados_graficos(chamados, uma_semana_atras, data_atual, "Uma-semana")
    hoje, dados_abertos_hoje, dados_fechados_hoje = dados_graficos(chamados,datetime.now().replace(hour=0, minute=0, second=0, microsecond=0), datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999), 'Hoje')
    
    tipos = TipoChamado.objects.all()

    dados_mes_tipo = {}
    dados_uma_semana_atras_tipo = {}
    dados_hoje_tipo = {}
    dados_este_ano_tipo = {}
    dados_um_ano_tipo = {}
    # Loop para processar cada tipo de chamado
    for tipo in tipos:
        tipo_id = str(tipo.id) 
        # informacoes por tipo
        # um mes atras
        semanas_tipo, dados_abertos_tipo, dados_fechados = dados_graficos_tipo(chamados, um_mes_atras, data_atual, "Um-mes", tipo_id)
        # uma semana atras
        semanas_tipo_uma_semana_atras, dados_abertos_tipo_uma_semana_atras, dados_fechados_uma_semana_atras = dados_graficos_tipo(chamados, uma_semana_atras, data_atual, "Uma-semana", tipo_id)
        #hoje
        semanas_tipo_hoje, dados_abertos_tipo_hoje, dados_fechados_hoje = dados_graficos_tipo(chamados, datetime.now().replace(hour=0, minute=0, second=0, microsecond=0), datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999), "Hoje", tipo_id)
        # este ano
        semanas_tipo_este_ano, dados_abertos_tipo_este_ano, dados_fechados_este_ano = dados_graficos_tipo(chamados,esse_ano , data_atual, "Este-ano", tipo_id)
        # um ano atras ate hoje
        semanas_tipo_um_ano, dados_abertos_tipo_um_ano, dados_fechados_um_ano = dados_graficos_tipo(chamados, um_ano , data_atual, "Um-ano", tipo_id)

        # Adiciona os dados ao dicionário
        dados_mes_tipo[tipo_id] = {
            'semanas': semanas_tipo,
            'dados_abertos': dados_abertos_tipo,
            'dados_fechados': dados_fechados,
        }
        dados_uma_semana_atras_tipo[tipo_id] = {
            'semanas': semanas_tipo_uma_semana_atras,
            'dados_abertos': dados_abertos_tipo_uma_semana_atras,
            'dados_fechados': dados_fechados_uma_semana_atras,
        }
        dados_hoje_tipo[tipo_id] = {
            'semanas': semanas_tipo_hoje,
            'dados_abertos': dados_abertos_tipo_hoje,
            'dados_fechados': dados_fechados_hoje,
        }
        dados_este_ano_tipo[tipo_id] = {
            'semanas': semanas_tipo_este_ano,
            'dados_abertos': dados_abertos_tipo_este_ano,
            'dados_fechados': dados_fechados_este_ano
        }
        dados_um_ano_tipo[tipo_id] = {
            'semanas': semanas_tipo_um_ano,
            'dados_abertos': dados_abertos_tipo_um_ano,
            'dados_fechados': dados_fechados_um_ano
        }

    # Chamados abertos há mais de 7 dias, ordenados por prioridade (alta primeiro), depois por tempo em aberto
    chamados_criticos_qs = Chamado.objects.filter(
        status='0',
        dt_inclusao__lte=timezone.now() - timedelta(days=7)
    ).order_by('-prioridade', 'dt_inclusao')

    chamados_criticos = []
    for chamado in chamados_criticos_qs:
        dias_em_aberto = (timezone.now().date() - chamado.dt_inclusao.date()).days
        chamado.dias_em_aberto = dias_em_aberto
        chamados_criticos.append(chamado)

    # Dividir chamados críticos em páginas de 9
    chamados_criticos_paginas = [
        chamados_criticos[i:i+9] for i in range(0, len(chamados_criticos), 9)
    ]

    context= {
        'dados_mes_tipo': dados_mes_tipo,
        'dados_uma_semana_atras_tipo': dados_uma_semana_atras_tipo,
        'dados_hoje_tipo': dados_hoje_tipo,
        'dados_este_ano_tipo': dados_este_ano_tipo,
        'dados_um_ano_tipo': dados_um_ano_tipo,
        'tipos': tipos,
        'count_abertos': count_abertos,
        'count_em_atendimento': count_em_atendimento,
        'count_pendentes': count_pendentes,
        'count_fechados': count_fechados,
        'count_finalizados': count_finalizados,
        'chamados_por_tipo': chamados_por_tipo,
        'dados_abertos_uma_semana': dados_abertos_uma_semana,
        'dados_fechados_uma_semana': dados_fechados_uma_semana,
        'uma_semana': uma_semana,
        'hoje': hoje,
        'dados_abertos_hoje': dados_abertos_hoje,
        'dados_fechados_hoje': dados_fechados_hoje,
        'um_mes_atras': semanas_mes,
        'dados_abertos_um_mes': dados_abertos_um_mes,
        'dados_fechados_um_mes': dados_fechados_um_mes,
        'meses': meses_ano,
        'dados_abertos_um_ano': dados_abertos_um_ano,
        'dados_fechados_um_ano': dados_fechados_um_ano,
        'meses_esse_ano': meses_esse_ano,
        'dados_abertos_esse_ano': dados_abertos_esse_ano,
        'dados_fechados_esse_ano': dados_fechados_esse_ano,
        'total_chamados': total_chamados,
        'chamados_abertos_30dias': chamados_abertos_30dias,
        'chamados_fechados_30dias': chamados_fechados_30dias,
        'chamados_finalizados_30dias': chamados_finalizados_30dias,
        'media_diaria': "{:.1f}".format(media_diaria),
        'chamados_abertos_por_tipo': chamados_abertos_por_tipo,
        'chamado': chamados.first(),
        'total_nao_atribuidos': total_nao_atribuidos,
        'porcentagem_nao_atribuidos': "{:.1f}".format(porcentagem_nao_atribuidos), 
        'variacao_30_a_60_dias': "{:.1f}".format(variacao_30_a_60_dias),
        'aumento_chamados_abertos_30_dias': aumento_chamados_abertos_30_dias,
        'total_nao_resolvidos':  total_nao_resolvidos,
        'porcentagem_nao_resolvidos': "{:.1f}".format(porcentagem_nao_resolvidos),
        'chamados_abertos_por_secretaria': chamados_abertos_por_secretaria,
        'total_chamados_abertos':total_chamados_abertos,
        'labels_atendimentos_realizados': labels_atendimentos_realizados,
        'dados_abertos_atendimentos_realizados': dados_abertos_atendimentos_realizados,
        'chamados_criticos': chamados_criticos,
        'chamados_criticos_paginas': chamados_criticos_paginas,
    }
    return render(request, 'chamados/new_dashboard.html', context)

@login_required
def ver_detalhes_tickets_nao_resolvidos(request):
    chamados_grupos = (
        TipoChamado.objects.annotate(
            aberto=Count('chamado', filter=Q(chamado__status='0')),
            pendente=Count('chamado', filter=Q(chamado__status='2')),
        )
        .annotate(total=Count('chamado'))
        .values('id','nome') 
        .annotate(aberto=Count('chamado', filter=Q(chamado__status='0')),
                  pendente=Count('chamado', filter=Q(chamado__status='2')),
                  total=Count('chamado')) 
        .order_by('nome') 
    )
     # Consulta para profissionais designados
    chamados_profissionais = (
        Atendente.objects.annotate(
            aberto=Count('profissional_designado_chamados', filter=Q(profissional_designado_chamados__status='0')),
            pendente=Count('profissional_designado_chamados', filter=Q(profissional_designado_chamados__status='2')),
            total=Count('profissional_designado_chamados')
        )
        .values('id', 'nome_servidor', 'aberto', 'pendente', 'total')
        .order_by('nome_servidor')
    )

    context = {
        'chamados_grupos': chamados_grupos,
        'chamados_profissionais': chamados_profissionais
    }

    return render(request, 'chamados/verDetalhesTicketsNaoResolvidos.html', context)

@login_required
def ver_perfil(request,matricula):
    is_atendente = Atendente.objects.filter(servidor=Servidor.objects.get(user=request.user)).exists(),
    servidor= get_object_or_404(Servidor, matricula=matricula)

    context = {
        'usuario': servidor,
        'tipos': TipoChamado.objects.all()
    }
    if is_atendente or servidor.user == request.user:
        chamado = Chamado.objects.filter(requisitante = servidor).order_by('-dt_inclusao')
        context['chamado']= chamado
    return render(request, 'chamados/perfil.html', context)

# VIEW PARA VERIFICAR ATUALIZAÇÃO OU CRIAÇÃO DOS CHAMADOS
@login_required
def verificar_atualizacoes(request):
    last_check = request.GET.get('last_check')

    if last_check:
        last_check = last_check.replace(' ', '+')
        try:
            last_check = timezone.datetime.fromisoformat(last_check)
        except ValueError as e:
            print(f"Erro ao converter a data: {e}")
            last_check = timezone.now() - timezone.timedelta(seconds=5)
    else:
        last_check = timezone.now() - timezone.timedelta(seconds=5)

    novos_chamados = Chamado.objects.filter(dt_inclusao__gt=last_check).count()
    atualizacoes = Chamado.objects.filter(dt_atualizacao__gt=last_check, dt_inclusao__lte=last_check).count()

    return JsonResponse({
        'novos_chamados': novos_chamados,
        'atualizacoes': atualizacoes,
        'last_check': timezone.now().isoformat()
    })


@login_required
@require_POST
def mesclar_chamados(request):
    try:
        atendente = Atendente.objects.get(servidor__user=request.user, ativo=True)
        if atendente.nivel not in ['0', '2']:
            return JsonResponse({'status': 403, 'message': 'Acesso negado!'})
    except Atendente.DoesNotExist:
        return JsonResponse({'status': 403, 'message': 'Acesso negado!'})

    ids = request.POST.getlist('chamados')
    if not ids:
        return JsonResponse({'status': 400, 'message': 'Nenhum chamado fornecido!'})

    ids = ids[0].split(',')
    chamados_mesclados = Chamado.objects.filter(id__in=ids).order_by('dt_inclusao')

    if not chamados_mesclados.exists():
        return JsonResponse({'status': 404, 'message': 'Chamados não encontrados!'})

    try:
        primeiro_chamado = chamados_mesclados.first()
        chamado_resultante = Chamado.objects.create(
            setor=primeiro_chamado.setor,
            secretaria=primeiro_chamado.secretaria,
            telefone=primeiro_chamado.telefone,
            requisitante=primeiro_chamado.requisitante,
            endereco=primeiro_chamado.endereco,
            tipo=primeiro_chamado.tipo,
            assunto=primeiro_chamado.assunto,
            prioridade=primeiro_chamado.prioridade,
            status='0',
            descricao=primeiro_chamado.descricao,
            profissional_designado=primeiro_chamado.profissional_designado,
            user_inclusao=Servidor.objects.get(user=request.user),
            anexo=primeiro_chamado.anexo,
            mesclado=True
        )
        chamado_resultante.gerar_hash()
        chamado_resultante.gerar_protocolo()

        if chamado_resultante.tipo.sigla == 'TEL':
            dados_telefonia = OSTelefonia.objects.filter(chamado=primeiro_chamado)
            if dados_telefonia.exists():
                OSTelefonia.objects.create(
                    chamado=chamado_resultante,
                    ramal=dados_telefonia.first().ramal,                    
                )


        for chamado in chamados_mesclados:
            Mensagem.objects.filter(chamado=chamado).update(chamado=chamado_resultante)
            chamado.mesclado = True
            chamado.status = '6'
            chamado.save()

        chamado_resultante.dt_atualizacao = timezone.now()
        chamado_resultante.save()

        return JsonResponse({'status': 200, 'message': 'Chamados mesclados com sucesso!'})

    except Exception as e:
        return JsonResponse({'status': 500, 'message': f'Erro interno: {str(e)}'})
    
def download_relatorio(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden('Você não tem autorização para acessar esta página!')
    
    chamados = filtrar_chamados(request)
    # print(chamados.count())
    chamados_ids = [chamado.id for chamado in chamados]
    # print(chamados_ids.count())
    chamados_queryset = Chamado.objects.filter(id__in=chamados_ids)    
    # print(chamados_queryset.count())
    # finalizados = chamados_queryset.filter(status='4')
    finalizados = chamados_queryset #gambiarra
    # print(finalizados.count())
    campos = ['n_protocolo', 
              'profissional_designado', 
              'tipo',
              'secretaria'
              'dt_inclusao',
              'dt_execucao',
              'dt_fechamento',
              'subtipo',
              'observacao'
            ]
    
    # Cria uma lista de dicionários com os dados dos chamados finalizados
    data = [
        {
            'Protocolo': chamado.n_protocolo,
            'Técnico': chamado.profissional_designado.nome_servidor if chamado.profissional_designado else '',
            'Grupos': chamado.tipo.nome,
            'Secretaria': chamado.secretaria.nome,
            'Data de Inclusão': chamado.dt_inclusao.strftime('%d/%m/%Y'),
            'Data de Execução': chamado.dt_execucao.strftime('%d/%m/%Y') if chamado.dt_execucao else '',
            'Data de Fechamento': chamado.dt_fechamento.strftime('%d/%m/%Y') if chamado.dt_fechamento else '',
            'Serviços': chamado.subtipo,
            'Observação': chamado.relatorio ,
        }
        for chamado in finalizados
    ]

    # Cria um DataFrame a partir dos dados
    df = pd.DataFrame(data)

    # Cria uma resposta HTTP com o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_chamados_finalizados.xlsx'

    # Salva o DataFrame no arquivo Excel
    df.to_excel(response, index=False)

    return response

def api_get_data_servidor(request):
    atendente = Atendente.objects.filter(servidor__user=request.user)
    if not atendente.exists():
        return JsonResponse({'status': 403, 'message': 'Acesso negado!'})
    elif not atendente.last().nivel in ['0', '2']:
            return JsonResponse({'status': 403, 'message': 'Acesso negado!'})
    
    if request.method == 'POST':
        id = request.POST.get('id')
        servidor = Servidor.objects.get(id=id)
        return JsonResponse({
            'status': 200,
            'secretaria_id': servidor.setor.secretaria.id,
            'secretaria': servidor.setor.secretaria.nome,
            'endereco': servidor.setor.endereco,            
            'telefone': servidor.telefone
        })

    return JsonResponse({'status': 400, 'message': 'Método não permitido!'})

from dashboards.graficos import dados_chamados_por_secretaria

def new_dashboard(request):
    
    
    graficos = [        
        {'id': 'chartA1', 'tipo': 'bar', 'dados': date_chamados_por_secretaria(), 'opcoes': options_chamados_por_secretaria(), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        {'id': 'chartC1', 'tipo': 'line', 'dados': data_generic('impressora'), 'opcoes': options_generic('impressora'), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        {'id': 'chartC2', 'tipo': 'line', 'dados': data_generic('internet'), 'opcoes': options_generic('internet'), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        {'id': 'chartC3', 'tipo': 'line', 'dados': data_generic('sistemas - E&L'), 'opcoes': options_generic('sistemas - E&L'), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        {'id': 'chartD1', 'tipo': 'line', 'dados': data_generic('computador'), 'opcoes': options_generic('computador'), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        {'id': 'chartD2', 'tipo': 'line', 'dados': data_generic('telefonia'), 'opcoes': options_generic('telefonia'), 'scales': {}, 'plugins': '[ChartDataLabels]'},

        {'id': 'chartA2', 'tipo': 'bar', 'dados': date_chamados_por_atendente(), 'opcoes': options_chamados_por_atendente(), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        {'id': 'chartA3', 'tipo': 'bar', 'dados': date_chamados_por_mes(), 'opcoes': options_chamados_por_mes(), 'scales': {}, 'plugins': '[ChartDataLabels]'},
        # {'id': 'chartD3', 'tipo': 'line', 'dados': date_generic(), 'opcoes': options_generic('a'), 'scales': {}, 'plugins': '[ChartDataLabels]'},

    ]
    # try:
    #     noticias = get_news()
    # except Exception as e:
    #     print(f"Erro ao buscar notícias: {e}")
    #     noticias = []

    # Chamados abertos há mais de 7 dias, ordenados por prioridade (alta primeiro), depois por tempo em aberto
    chamados_criticos_qs = Chamado.objects.filter(
        status='0',
        dt_inclusao__lte=timezone.now() - timedelta(days=7)
    ).order_by('-prioridade', 'dt_inclusao')

    chamados_criticos = []
    for chamado in chamados_criticos_qs:
        dias_em_aberto = (timezone.now().date() - chamado.dt_inclusao.date()).days
        chamado.dias_em_aberto = dias_em_aberto
        chamados_criticos.append(chamado)

    # Dividir chamados críticos em páginas de 9
    chamados_criticos_paginas = [
        chamados_criticos[i:i+9] for i in range(0, len(chamados_criticos), 9)
    ]

    context= {
        'graficos': graficos,
        'totais': {
            'total_chamados': Chamado.objects.filter(~Q(status__in=['5', '6']), dt_inclusao__gte=timezone.now() - timedelta(days=30)).count(),
            'chamados_abertos': Chamado.objects.filter(status='0', dt_inclusao__gte=timezone.now() - timedelta(days=30)).count(),
            'chamados_pendentes': Chamado.objects.filter(status='2', dt_inclusao__gte=timezone.now() - timedelta(days=30)).count(),
            'chamados_fechados': Chamado.objects.filter(status='4', dt_inclusao__gte=timezone.now() - timedelta(days=30)).count(),            
        },
        'taxa_eficiencia': round(Chamado.objects.filter(status='4', dt_inclusao__gte=timezone.now() - timedelta(days=30)).count() / Chamado.objects.filter(~Q(status__in=['5', '6']), dt_inclusao__gte=timezone.now() - timedelta(days=30)).count() * 100, 1) if Chamado.objects.filter(~Q(status__in=['5', '6']), dt_inclusao__gte=timezone.now() - timedelta(days=30)).count() > 0 else 0,
        'media_diaria': round(Chamado.objects.filter(~Q(status__in=['6']), dt_inclusao__gte=timezone.now() - timedelta(days=30)).count() / 30, 1),
        # 'noticias': noticias
        'chamados_criticos': chamados_criticos,
        'chamados_criticos_paginas': chamados_criticos_paginas,
    }
    return render(request, 'chamados/new_dashboard.html', context)


@login_required
def new_dashboard_2(request):
    from .models import Chamado
    from datetime import timedelta
    from django.utils import timezone

    # Chamados abertos há mais de 7 dias, ordenados por prioridade (alta primeiro), depois por tempo em aberto
    chamados_criticos_qs = Chamado.objects.filter(
        status='0',
        dt_inclusao__lte=timezone.now() - timedelta(days=7)
    ).order_by('-prioridade', 'dt_inclusao')

    chamados_criticos = []
    for chamado in chamados_criticos_qs:
        dias_em_aberto = (timezone.now().date() - chamado.dt_inclusao.date()).days
        chamado.dias_em_aberto = dias_em_aberto
        chamados_criticos.append(chamado)

    # Dividir chamados críticos em páginas de 9
    chamados_criticos_paginas = [
        chamados_criticos[i:i+9] for i in range(0, len(chamados_criticos), 9)
    ]

    context = {
        'chamados_criticos': chamados_criticos,
        'chamados_criticos_paginas': chamados_criticos_paginas,
    }
    return render(request, 'chamados/new_dashboard_2.html', context)


import requests
def get_news():
    API_KEY = 'a562b55b3c794fff85f3492b3a72fa15'
    API_URL = f'https://newsapi.org/v2/everything?q="pt-br"&apiKey={API_KEY}'

    
    try:
        response = requests.get(API_URL)
        response.raise_for_status()  # Levanta uma exceção para códigos de status HTTP de erro
        data = response.json()
        if data['status'] == 'ok':
            return data['articles'][:6]  # Retorna apenas as 5 primeiras notícias
        return []
    except requests.exceptions.RequestException as e:
        print(f"Erro ao buscar notícias: {e}")
        return []
    from django.shortcuts import render
from django.db.models import Count
from .models import chamadoSatisfacao, Chamado
from colorsys import rgb_to_hls, hls_to_rgb

def ajustar_tom(cor_hex, fator):
    """Ajusta o tom de uma cor em formato hexadecimal."""
    cor_hex = cor_hex.lstrip('#')
    r, g, b = tuple(int(cor_hex[i:i+2], 16) for i in (0, 2, 4))
    r, g, b = [x / 255.0 for x in (r, g, b)]
    h, l, s = rgb_to_hls(r, g, b)
    l = max(0, min(1, l * fator))  # Ajusta a luminosidade
    r, g, b = hls_to_rgb(h, l, s)
    return f'#{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}'

def painel_satisfacao(request):
    feedbacks = chamadoSatisfacao.objects.filter(chamado__pesquisa_satisfacao=True).order_by('-id')
    
    totais = {
        'geral': sum(int(feed.avaliacao) for feed in feedbacks),
        'cordialidade': sum(int(feed.cordialidade) for feed in feedbacks),
        'negativos_geral': sum(1 for feed in feedbacks if int(feed.avaliacao) < 2),
        'negativos_cordialidade': sum(1 for feed in feedbacks if int(feed.cordialidade) < 2),
    }

    feedbacks_negativos = [feed for feed in feedbacks if int(feed.avaliacao) < 2 or int(feed.cordialidade) < 2]
    feedbacks_positivos = [feed for feed in feedbacks if feed not in feedbacks_negativos]

    def process_group_data(queryset, field):
        grouped_data = {}
        for dado in queryset:
            nome = dado['chamado__profissional_designado__nome_servidor'].split()[0]
            valor = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(dado[field])
            grouped_data.setdefault(nome, {label: 0 for label in avaliacao_labels})
            grouped_data[nome][valor] = dado['total']
        return grouped_data

    avaliacao_labels = list(dict(chamadoSatisfacao.AVALIACAO_CHOICES).values())
    atendente_labels = []

    avaliacao_atendentes = chamadoSatisfacao.objects.values(
        'chamado__profissional_designado__nome_servidor', 'avaliacao'
    ).annotate(total=Count('avaliacao')).order_by('chamado__profissional_designado__nome_servidor', 'avaliacao')

    cordialidade_atendentes = chamadoSatisfacao.objects.values(
        'chamado__profissional_designado__nome_servidor', 'cordialidade'
    ).annotate(total=Count('cordialidade')).order_by('chamado__profissional_designado__nome_servidor', 'cordialidade')

    data_avaliacao_por_atendente = process_group_data(avaliacao_atendentes, 'avaliacao')
    data_cordialidade_por_atendente = process_group_data(cordialidade_atendentes, 'cordialidade')

    atendente_labels = list(data_avaliacao_por_atendente.keys())

    cores = {
        'Excelente': '#007AFF',
        'Ótimo': '#34C759',
        'Bom': '#FFD60A',
        'Regular': '#FF9500',
        'Ruim': '#FF3B30',
    }

    def create_datasets(data_by_atendente):
        return [
            {
                'label': avaliacao,
                'data': [data_by_atendente[atendente][avaliacao] for atendente in atendente_labels],
                'backgroundColor': cores[avaliacao],
                'borderColor': cores[avaliacao],
                'borderWidth': 1,
            }
            for avaliacao in avaliacao_labels
        ]

    datasets_avaliacao = create_datasets(data_avaliacao_por_atendente)
    datasets_cordialidade = create_datasets(data_cordialidade_por_atendente)

    def aggregate_general_data(queryset, field):
        aggregated = {label: 0 for label in avaliacao_labels}
        for dado in queryset:
            label = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(dado[field])
            aggregated[label] += dado['total']
        return aggregated

    avaliacao_geral = chamadoSatisfacao.objects.values('avaliacao').annotate(total=Count('avaliacao'))
    cordialidade_geral = chamadoSatisfacao.objects.values('cordialidade').annotate(total=Count('cordialidade'))

    data_geral_avaliacao = aggregate_general_data(avaliacao_geral, 'avaliacao')
    data_geral_cordialidade = aggregate_general_data(cordialidade_geral, 'cordialidade')

    dataset_geral = lambda label, data: {
        'label': label,
        'data': [data[label] for label in avaliacao_labels],
        'backgroundColor': [cores[label] for label in avaliacao_labels],
        'borderColor': [cores[label] for label in avaliacao_labels],
        'borderWidth': 1,
    }

    context = {
        'feedbacks_positivos': feedbacks_positivos[:15],
        'feedbacks_positivos_count': len(feedbacks_positivos),
        'feedbacks_negativos': feedbacks_negativos[:15],
        'feedbacks_negativos_count': len(feedbacks_negativos),
        'media_geral': round(totais['geral'] / len(feedbacks), 1) if feedbacks else 0,
        'media_cordialidade': round(totais['cordialidade'] / len(feedbacks), 1) if feedbacks else 0,
        'qnt_negativos_geral': totais['negativos_geral'],
        'qnt_negativos_cordialidade': totais['negativos_cordialidade'],
        'feed_pendentes': Chamado.objects.filter(status='4', pesquisa_satisfacao=False).order_by('-id'),
        'atendente_labels': atendente_labels,
        'datasets_avaliacao': datasets_avaliacao,
        'datasets_cordialidade': datasets_cordialidade,
        'labels_geral': avaliacao_labels,
        'datasets_geral': [
            dataset_geral('Avaliação', data_geral_avaliacao),
            dataset_geral('Cordialidade', data_geral_cordialidade),
        ],
    }

    return render(request, 'chamados/painel_satisfacao.html', context)

def pesquisar_feedback(request):
    if request.method == 'GET':
        protocolo = request.GET.get('protocolo')
        if not protocolo:
            return JsonResponse({'error': 'Protocolo não fornecido'}, status=400)
        try:
            feedback = chamadoSatisfacao.objects.filter(chamado__n_protocolo__icontains=protocolo)
            if not feedback.exists():
                return JsonResponse({'status':404, 'error': 'Feedback não encontrado para o protocolo fornecido.'}, status=200)
            feedback = feedback.first()
            response_data = {
                'status': 200,
                'chamado': {
                            'n_protocolo': feedback.chamado.n_protocolo,
                            'secretaria': feedback.chamado.secretaria.nome,
                            'telefone': feedback.chamado.telefone,
                            'requisitante': feedback.chamado.requisitante.nome,
                            'profissional_designado': feedback.chamado.profissional_designado.nome_servidor if feedback.chamado.profissional_designado else 'N/H',
                            'descricao': feedback.chamado.descricao,
                            'tipo': feedback.chamado.tipo.nome,
                            'subtipo': feedback.chamado.subtipo.nome,
                            'relatorio': feedback.chamado.relatorio,
                            'link': reverse('chamados:detalhes', args=[feedback.chamado.hash]),
                            'data_abertura': feedback.chamado.dt_inclusao.strftime('%d/%m/%Y'),
                            },
                'avaliacao_geral': feedback.get_avaliacao_display(),
                'geral_justificativa': feedback.avaliacao_justificativa if feedback.avaliacao_justificativa else 'N/H',
                'avaliacao_cordialidade': feedback.get_cordialidade_display(),
                'cordialidade_justificativa': feedback.cordialidade_justificativa if feedback.cordialidade_justificativa else 'N/H',
                'resolucao': feedback.get_resolucao_display(),
                'receberia_novamente': feedback.get_receberia_novamente_o_tecnico_display(),
                'tempo_espera': feedback.get_tempo_espera_display(),
                'sugestoes': feedback.comentario if feedback.comentario != '' else 'N/H',
                
                'data_feedback': feedback.dt_inclusao.strftime('%d/%m/%Y'),
            }
            return JsonResponse(response_data, status=200)
        except Exception as e:
            return JsonResponse({'status': 500, 'error': str(e)}, status=200)
    return JsonResponse({'status':405,'error': 'Método não permitido.'}, status=405)


# def feedback_in_excel(request):
#     # Criação do arquivo Excel
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     hoje = datetime.now().strftime('%d/%m/%Y_%H:%M:%S')
#     sheet.title = "Feedbacks"

#     # Cabeçalhos
#     headers = [
#         "Chamado",
#         "Técnico",
#         "Avaliação",
#         "Justificativa da Avaliação",
#         "Cordialidade",
#         "Justificativa da Cordialidade",
#         "Resolução",
#         "Receberia Novamente o Técnico",
#         "Tempo de Espera",
#         "Comentário",
#         "Data de Inclusão"
#     ]
    
#     # Estilo para o cabeçalho
#     header_font = Font(bold=True)
    
#     for col_num, header in enumerate(headers, 1):
#         cell = sheet.cell(row=1, column=col_num)
#         cell.value = header
#         cell.font = header_font

#     # Dados dos feedbacks
#     feedbacks = chamadoSatisfacao.objects.all()
#     for row_num, feedback in enumerate(feedbacks, 2):
#         sheet.cell(row=row_num, column=1).value = feedback.chamado.n_protocolo
#         sheet.cell(row=row_num, column=2).value = feedback.chamado.profissional_designado.nome_servidor if feedback.chamado.profissional_designado else "N/H"
#         sheet.cell(row=row_num, column=3).value = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(feedback.avaliacao, "N/A")
#         sheet.cell(row=row_num, column=4).value = feedback.avaliacao_justificativa
#         sheet.cell(row=row_num, column=5).value = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(feedback.cordialidade, "N/A")
#         sheet.cell(row=row_num, column=6).value = feedback.cordialidade_justificativa
#         sheet.cell(row=row_num, column=7).value = dict(chamadoSatisfacao.RESOLUCAO_CHOICES).get(feedback.resolucao, "N/A")
#         sheet.cell(row=row_num, column=8).value = dict(chamadoSatisfacao.RECEBER_CHOICES).get(feedback.receberia_novamente_o_tecnico, "N/A")
#         sheet.cell(row=row_num, column=9).value = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(feedback.tempo_espera, "N/A")
#         sheet.cell(row=row_num, column=10).value = feedback.comentario
#         sheet.cell(row=row_num, column=11).value = feedback.dt_inclusao.strftime('%d/%m/%Y %H:%M:%S')

#     # Ajuste de largura das colunas (opcional)
#     for col in sheet.columns:
#         max_length = 0
#         col_letter = col[0].column_letter  # Obter letra da coluna
#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         adjusted_width = max_length + 2
#         sheet.column_dimensions[col_letter].width = adjusted_width

#     # Retorno como arquivo de download
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="feedbacks_{hoje}.xlsx"'
#     workbook.save(response)
#     return response

def feedback_in_excel(request):
    # Criação do arquivo Excel
    workbook = openpyxl.Workbook()
    hoje = datetime.now().strftime('%d_%m_%Y_%H_%M_%S')
    
    # Aba principal
    sheet_feedbacks = workbook.active
    sheet_feedbacks.title = "Feedbacks"
    
    # Aba adicional
    aba_pendentes = workbook.create_sheet("Pendentes")
    aba_resumo = workbook.create_sheet("Resumo")
    
    # Cabeçalhos para Feedbacks
    headers = [
        "Chamado", "Técnico", "Avaliação", "Justificativa da Avaliação", 
        "Cordialidade", "Justificativa da Cordialidade", "Resolução", 
        "Receberia Novamente o Técnico", "Tempo de Espera", "Comentário", "Data de Inclusão"
    ]
    
    # Estilo para cabeçalhos
    header_font = Font(bold=True)
    
    # Preenche cabeçalhos na aba principal
    for col_num, header in enumerate(headers, 1):
        cell = sheet_feedbacks.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
    
    # Dados dos feedbacks
    chamados = Chamado.objects.all()    
    total_chamados_abertos_ou_pendentes = chamados.filter(status__in=['0', '2']).count()
    total_chamados_fechados = chamados.filter(status='4').count()
    total_chamados_cancelados_ou_mesclados = chamados.filter(status__in=['5', '6']).count()
    total_dias = (timezone.now() - chamados.first().dt_inclusao).days    
    total_dias = total_dias if total_dias > 0 else 1    
    media_diaria_geral = round(chamados.count() / total_dias, 1)
    feedbacks = chamadoSatisfacao.objects.all()
    feed_pendente = Chamado.objects.filter(status='4', pesquisa_satisfacao=False).order_by('-id')
    
    total_feedbacks = feedbacks.count()
    total_feedbacks_positivos = feedbacks.filter(avaliacao__gte=2, cordialidade__gte=2).count()
    total_feedbacks_negativos = feedbacks.filter(Q(avaliacao__lt=2) | Q(cordialidade__lt=2)).count()
    total_feedbacks_pendentes = feed_pendente.count()
    
    chamados_ultimos30dias = chamados.filter(dt_inclusao__gte=timezone.now() - timedelta(days=30))
    chamados_abertos_ou_pendentes_30dias = chamados_ultimos30dias.filter(status__in=['0', '2']).count()
    chamados_fechados_30dias = chamados_ultimos30dias.filter(status='4').count()
    chamados_cancelados_ou_mesclados_30dias = chamados_ultimos30dias.filter(status__in=['5', '6']).count()    
    media_diaria_30dias = round(chamados_ultimos30dias.count() / 30, 1)

    for row_num, feedback in enumerate(feedbacks, 2):
        sheet_feedbacks.cell(row=row_num, column=1).value = feedback.chamado.n_protocolo
        sheet_feedbacks.cell(row=row_num, column=2).value = feedback.chamado.profissional_designado.nome_servidor if feedback.chamado.profissional_designado else "N/H"
        sheet_feedbacks.cell(row=row_num, column=3).value = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(feedback.avaliacao, "N/A")
        sheet_feedbacks.cell(row=row_num, column=4).value = feedback.avaliacao_justificativa
        sheet_feedbacks.cell(row=row_num, column=5).value = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(feedback.cordialidade, "N/A")
        sheet_feedbacks.cell(row=row_num, column=6).value = feedback.cordialidade_justificativa
        sheet_feedbacks.cell(row=row_num, column=7).value = dict(chamadoSatisfacao.RESOLUCAO_CHOICES).get(feedback.resolucao, "N/A")
        sheet_feedbacks.cell(row=row_num, column=8).value = dict(chamadoSatisfacao.RECEBER_CHOICES).get(feedback.receberia_novamente_o_tecnico, "N/A")
        sheet_feedbacks.cell(row=row_num, column=9).value = dict(chamadoSatisfacao.AVALIACAO_CHOICES).get(feedback.tempo_espera, "N/A")
        sheet_feedbacks.cell(row=row_num, column=10).value = feedback.comentario
        sheet_feedbacks.cell(row=row_num, column=11).value = feedback.dt_inclusao.strftime('%d/%m/%Y %H:%M:%S')

    for row_num, feed in enumerate(feed_pendente, 2):
        #n_protocolo, requisitante, profissional_designado, tipo, dt_inclusao, dt_execucao, dt_fechamento
        aba_pendentes.cell(row=row_num, column=1).value = feed.n_protocolo
        aba_pendentes.cell(row=row_num, column=2).value = feed.requisitante.nome
        aba_pendentes.cell(row=row_num, column=3).value = feed.profissional_designado.nome_servidor if feed.profissional_designado else "N/H"
        aba_pendentes.cell(row=row_num, column=4).value = feed.tipo.nome
        aba_pendentes.cell(row=row_num, column=5).value = feed.dt_inclusao.strftime('%d/%m/%Y %H:%M:%S')
        aba_pendentes.cell(row=row_num, column=6).value = feed.dt_execucao.strftime('%d/%m/%Y %H:%M:%S') if feed.dt_execucao else ''
        aba_pendentes.cell(row=row_num, column=7).value = feed.dt_fechamento.strftime('%d/%m/%Y %H:%M:%S') if feed.dt_fechamento else ''
        
    # Preenche conteúdo da aba adicional
    aba_resumo.cell(row=1, column=1).value = "Feedbacks"
    aba_resumo.cell(row=1, column=1).font = header_font

    aba_resumo.cell(row=2, column=1).value = "Total de Feedbacks"
    aba_resumo.cell(row=2, column=2).value = total_feedbacks
    aba_resumo.cell(row=3, column=1).value = "Total de Feedbacks Positivos"
    aba_resumo.cell(row=3, column=2).value = total_feedbacks_positivos
    aba_resumo.cell(row=4, column=1).value = "Total de Feedbacks Negativos"
    aba_resumo.cell(row=4, column=2).value = total_feedbacks_negativos
    aba_resumo.cell(row=5, column=1).value = "Total de Feedbacks Pendentes"
    aba_resumo.cell(row=5, column=2).value = total_feedbacks_pendentes

    aba_resumo.cell(row=7, column=1).value = "Chamados nos Últimos 30 Dias"
    aba_resumo.cell(row=7, column=1).font = header_font

    aba_resumo.cell(row=8, column=1).value = "Total de Chamados"
    aba_resumo.cell(row=8, column=2).value = chamados_ultimos30dias.count()
    aba_resumo.cell(row=9, column=1).value = "Abertos ou Pendentes"
    aba_resumo.cell(row=9, column=2).value = chamados_abertos_ou_pendentes_30dias
    aba_resumo.cell(row=10, column=1).value = "Fechados"
    aba_resumo.cell(row=10, column=2).value = chamados_fechados_30dias
    aba_resumo.cell(row=11, column=1).value = "Cancelados ou Mesclados"
    aba_resumo.cell(row=11, column=2).value = chamados_cancelados_ou_mesclados_30dias
    aba_resumo.cell(row=12, column=1).value = "Média Diária"
    aba_resumo.cell(row=12, column=2).value = media_diaria_30dias

    aba_resumo.cell(row=14, column=1).value = "Chamados Totais"
    aba_resumo.cell(row=14, column=1).font = header_font
    aba_resumo.cell(row=15, column=1).value = "Total de Chamados"
    aba_resumo.cell(row=15, column=2).value = chamados.count()
    aba_resumo.cell(row=16, column=1).value = "Abertos ou Pendentes"
    aba_resumo.cell(row=16, column=2).value = total_chamados_abertos_ou_pendentes
    aba_resumo.cell(row=17, column=1).value = "Fechados"
    aba_resumo.cell(row=17, column=2).value = total_chamados_fechados
    aba_resumo.cell(row=18, column=1).value = "Cancelados ou Mesclados"
    aba_resumo.cell(row=18, column=2).value = total_chamados_cancelados_ou_mesclados
    aba_resumo.cell(row=19, column=1).value = "Média Diária"
    aba_resumo.cell(row=19, column=2).value = media_diaria_geral
    # Ajuste opcional da largura das colunas
    for col in sheet_feedbacks.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet_feedbacks.column_dimensions[col_letter].width = adjusted_width
    
    # Retorno como arquivo para download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="feedbacks_{hoje}.xlsx"'
    workbook.save(response)
    return response

# detalhes = relatorio do chamado
def editar_detalhes_chamado(request, hash):
    instance = Chamado.objects.filter(hash=hash)
    
    if not instance.exists():
            return HttpResponseNotFound('Chamado não encontrado!')
    instance = instance.first()
    
    if not request.user.is_superuser:               
        if instance.profissional_designado != Atendente.objects.filter(servidor__user=request.user).first():
            return HttpResponseForbidden('Você não tem permissão para editar este relatório!')
        if instance.status != 4:
            return HttpResponseForbidden('Apenas chamados finalizados podem ser editados!')
        
    if request.method == 'POST':
        form = FormDetalhesDoChamado(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('chamados:detalhes', hash=hash)        
    else:
        form = FormDetalhesDoChamado(instance=instance)
    
    context = {
                'form': form,
                'chamado': instance
              }
    return render(request, 'chamados/editar_detalhes_chamado.html', context)